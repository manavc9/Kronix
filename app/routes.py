from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, flash
from .models import db, Location, Spot, YardTrack, Customer, User, Car, WorkOrder, \
                    SpotAssignment, YardCar, CleaningEntry, ScheduleEntry, Comment, ActivityLog, \
                    MonthlyTarget, GanttMark
from datetime import datetime, date, timedelta
from collections import OrderedDict, defaultdict
import calendar
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

main = Blueprint('main', __name__)


# ─────────────────────────────────────────────
# GANTT CONSTANTS
# ─────────────────────────────────────────────

LOCATION_CAPACITY = {
    'Thornfield': 780,
    'Delverton':  520,
    'Harwick':    340,
}

GANTT_SPOT_PATTERNS = {
    'Thornfield': ['Bay-A-', 'Bay-B-', 'SP-'],
    'Delverton':  ['Bay-C-', 'Bay-D-'],
    'Harwick':    ['Bay-E-', 'Bay-F-'],
}


def is_gantt_spot(spot_name, location_name):
    """Returns True if the spot should appear on the Gantt view."""
    patterns = GANTT_SPOT_PATTERNS.get(location_name, [])
    return any(spot_name.startswith(p) or spot_name == p for p in patterns)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def get_current_user():
    """Get the logged-in user from session."""
    user_id = session.get('user_id')
    if user_id:
        return User.query.get(user_id)
    return None

def log_activity(work_order_id, user, role, action, field=None, old_val=None, new_val=None):
    """Record every change to the activity log."""
    entry = ActivityLog(
        work_order_id=work_order_id,
        user=user,
        role=role,
        action=action,
        field_name=field,
        old_value=str(old_val) if old_val is not None else None,
        new_value=str(new_val) if new_val is not None else None,
    )
    db.session.add(entry)

def get_status_color(status):
    """Returns a badge color class for each status."""
    colors = {
        # Pre-repair
        'To Arrive':                   'gray',
        'Arrived':                     'blue',
        'Waiting Inspection':          'amber',
        'Waiting Estimate':            'amber',
        'Waiting Approval':            'amber',
        'Waiting Work Order':          'gray',
        'Work Order Created':          'teal',
        'Inbounded':                   'blue',
        # Cleaning
        'Waiting Cleaning':            'amber',
        'Scheduled':                   'blue',
        'Ready for Cleaning':          'amber',
        'In Cleaning':                 'accent',
        'Cleaned':                     'teal',
        # Repair
        'Waiting Repair':              'purple',
        'Waiting Material':            'red',
        'In Repair':                   'accent',
        'Under Repair':                'accent',
        'Repair Complete':             'green',
        # Post-repair
        'Waiting Outbound Inspection': 'amber',
        'Waiting Dispo':               'amber',
        'Dispo Received':              'teal',
        'Ready to Bill':               'green',
        'Billed':                      'green',
        'Completed':                   'green',
        # Special
        'On Hold':                     'red',
        'Cancelled':                   'gray',
        'Rework Needed':               'red',
        'Waiting Customer Approval':   'amber',
        # Yard
        'Just Inbounded':              'blue',
        'In Yard':                     'blue',
        'In Yard Repair':              'accent',
    }
    return colors.get(status, 'gray')


@main.route('/api/search-cars')
def search_cars_api():
    user = get_current_user()
    if not user:
        return jsonify([])
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    wos = WorkOrder.query.join(Car).filter(
        WorkOrder.is_closed == False,
        db.or_(
            Car.car_mark.ilike(f'%{q}%'),
            Car.car_number.ilike(f'%{q}%'),
        )
    ).limit(8).all()
    return jsonify([{
        'wo_id':    wo.id,
        'full_id':  wo.car.full_id,
        'customer': wo.car.customer.name if wo.car.customer else '—',
        'status':   wo.status,
    } for wo in wos])


def auto_calculate_status(wo):
    """Infer WO status from filled date fields (most-advanced milestone wins)."""
    if wo.final_estimate_approved:
        inferred = 'Completed'
    elif wo.billed_out:
        inferred = 'Billed'
    elif wo.dispo_received:
        inferred = 'Dispo Received'
    elif wo.dispo_requested or wo.outbound_inspection:
        inferred = 'Waiting Dispo'
    elif wo.repair_completed_date:
        inferred = 'Repair Complete'
    elif wo.repair_start_date:
        inferred = 'In Repair'
    elif wo.cleaning_completed:
        inferred = 'Cleaned'
    elif wo.cleaning_scheduled:
        inferred = 'Waiting Cleaning'
    elif wo.work_order_created:
        inferred = 'Work Order Created'
    elif wo.estimate_approved:
        inferred = 'Waiting Work Order'
    elif wo.estimate_sent:
        inferred = 'Waiting Approval'
    elif wo.inbound_inspection:
        inferred = 'Waiting Estimate'
    elif wo.date_arrived:
        inferred = 'Arrived'
    else:
        return
    if inferred != wo.status:
        wo.status = inferred


def get_scheduling_warnings(wo):
    """Return list of scheduling warning dicts for a work order."""
    warnings = []
    if wo.cleaning_required and not wo.cleaning_completed:
        warnings.append({'level': 'red', 'message': 'Cleaning required but not completed'})
    if wo.parts_status in ('Unknown', 'Ordered', 'Waiting ETA'):
        eta_str = f' — ETA {wo.parts_eta}' if wo.parts_eta else ''
        warnings.append({'level': 'orange', 'message': f'Parts: {wo.parts_status}{eta_str}'})
    if not wo.work_order_created:
        warnings.append({'level': 'yellow', 'message': 'Work order not yet created'})
    return warnings


def get_spot_zone(spot_name, location_name):
    """Map a spot name to its zone label based on location."""
    if location_name == 'Thornfield':
        if spot_name.startswith('CL-'):
            return 'Cleaning Rack'
        elif spot_name.startswith('Bay-A-') or spot_name.startswith('Bay-B-'):
            return 'Main Repair'
        elif spot_name.startswith('SP-'):
            return 'Surface Prep'
        return 'Other'
    elif location_name == 'Delverton':
        if spot_name.startswith('CL-'):
            return 'Cleaning Rack'
        elif spot_name.startswith('RW-'):
            return 'Rework Lane'
        return 'Main Repair'
    elif location_name == 'Harwick':
        if spot_name.startswith('CL-'):
            return 'Cleaning Rack'
        elif spot_name.startswith('Bay-E-'):
            return 'Indoor Repair'
        elif spot_name.startswith('Bay-F-'):
            return 'Outdoor Repair'
        return 'Other'
    return 'Other'


# ─────────────────────────────────────────────
# EXPORT HELPERS
# ─────────────────────────────────────────────

def style_excel_workbook(wb, title, subtitle=''):
    """Return a dict of Kronix-branded openpyxl style objects."""
    KRONOS_BLUE = '2563EB'
    WHITE       = 'FFFFFF'
    LIGHT_GRAY  = 'F1F5F9'
    MED_GRAY    = 'E2E8F0'
    DARK_TEXT   = '1E293B'

    header_fill  = PatternFill('solid', fgColor=KRONOS_BLUE)
    header_font  = Font(bold=True, color=WHITE, name='Calibri', size=11)
    header_align = Alignment(horizontal='center', vertical='center')
    title_font   = Font(bold=True, color=DARK_TEXT, name='Calibri', size=14)
    thin         = Side(style='thin', color='CBD5E1')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        'header_fill':  header_fill,
        'header_font':  header_font,
        'header_align': header_align,
        'title_font':   title_font,
        'border':       border,
        'LIGHT_GRAY':   LIGHT_GRAY,
        'MED_GRAY':     MED_GRAY,
        'WHITE':        WHITE,
    }


def excel_response(wb, filename):
    """Return an Excel file as a Flask download response."""
    from flask import Response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'},
    )


def pdf_response(buffer, filename):
    """Return a PDF file as a Flask download response."""
    from flask import Response
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}.pdf'},
    )


def _apply_xl_header_row(ws, headers, row_num, styles):
    """Write a styled header row to worksheet ws at row_num."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.fill      = styles['header_fill']
        cell.font      = styles['header_font']
        cell.alignment = styles['header_align']
        cell.border    = styles['border']
    ws.row_dimensions[row_num].height = 22


def _xl_title_block(ws, title_text, subtitle_text, col_span=10):
    """Write branded title + subtitle in rows 1-2, stats placeholder in row 4."""
    ws.merge_cells(f'A1:{get_column_letter(col_span)}1')
    ws['A1'] = title_text
    ws['A1'].font = Font(bold=True, size=16, color='2563EB', name='Calibri')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(col_span)}2')
    ws['A2'] = subtitle_text
    ws['A2'].font = Font(size=12, color='64748B', name='Calibri')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8   # visual spacer


def _xl_data_row(ws, row, data, styles):
    """Write a zebra-striped data row."""
    fill = PatternFill('solid', fgColor=styles['LIGHT_GRAY'] if row % 2 == 0 else styles['WHITE'])
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill      = fill
        cell.border    = styles['border']
        cell.alignment = Alignment(vertical='center')


def _pdf_title_elements(title_text, subtitle_text, elements):
    """Append Kronix-branded title + subtitle paragraphs."""
    title_style = ParagraphStyle('KronixTitle',
        fontName='Helvetica-Bold', fontSize=18,
        textColor=colors.HexColor('#2563EB'), spaceAfter=4)
    sub_style = ParagraphStyle('KronixSub',
        fontName='Helvetica', fontSize=11,
        textColor=colors.HexColor('#64748B'), spaceAfter=14)
    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(subtitle_text, sub_style))


def _pdf_section_header(label, elements):
    """Append a section divider bar paragraph."""
    style = ParagraphStyle('SectionHdr',
        fontName='Helvetica-Bold', fontSize=11,
        textColor=colors.white,
        backColor=colors.HexColor('#1E293B'),
        spaceBefore=10, spaceAfter=6,
        leftPadding=8, rightPadding=8,
        topPadding=4, bottomPadding=4)
    elements.append(Paragraph(label, style))


def _pdf_table_style(n_rows, overdue_rows=None):
    """Return a TableStyle for Kronix-branded data tables."""
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('ROWHEIGHT',  (0, 0), (-1, -1), 18),
    ]
    for i in range(1, n_rows):
        if overdue_rows and i in overdue_rows:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FEF2F2')))
        elif i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    return TableStyle(style)


def _pdf_footer(elements):
    """Append a standard Kronix footer."""
    elements.append(Spacer(1, 0.15 * inch))
    footer_style = ParagraphStyle('Footer',
        fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#94A3B8'))
    elements.append(Paragraph(
        f'Generated by Kronix Rail Operations Intelligence · '
        f'{date.today().strftime("%B %d, %Y")}',
        footer_style))


# ─────────────────────────────────────────────
# ERROR HANDLERS
# ─────────────────────────────────────────────

@main.app_errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@main.app_errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500


# ─────────────────────────────────────────────
# LOGIN / LOGOUT
# ─────────────────────────────────────────────

@main.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        user = User.query.get(user_id)
        if user:
            session['user_id']   = user.id
            session['user_name'] = user.name
            session['user_role'] = user.role
            session['location_id'] = user.location_id
            return redirect(url_for('main.dashboard'))
    users = User.query.filter_by(is_active=True).order_by(User.role, User.name).all()
    return render_template('login.html', users=users)

@main.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.login'))


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@main.route('/dashboard')
def dashboard():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    # Location filter — scheduler/admin see all, others see their location
    if user.location_id:
        loc_filter = [user.location_id]
    else:
        loc_filter = [l.id for l in Location.query.all()]

    # Active work orders for this user's scope
    active_wos = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.is_closed == False
    ).all()
    active_wos.sort(key=lambda w: w.days_on_site or 0, reverse=True)

    # Dashboard stats
    stats = {
        'total_active':    len(active_wos),
        'in_repair':       sum(1 for w in active_wos if w.status == 'In Repair'),
        'overdue':         sum(1 for w in active_wos if w.is_overdue),
        'waiting_approval':sum(1 for w in active_wos if w.status == 'Waiting Approval'),
        'waiting_material':sum(1 for w in active_wos if w.status == 'Waiting Material'),
        'waiting_cleaning':sum(1 for w in active_wos if w.status == 'Waiting Cleaning'),
        'completed_month': WorkOrder.query.filter(
            WorkOrder.location_id.in_(loc_filter),
            WorkOrder.status == 'Completed',
            WorkOrder.repair_completed_date >= date(date.today().year, date.today().month, 1)
        ).count(),
    }

    # Overdue cars — most critical first
    overdue_wos = [w for w in active_wos if w.is_overdue]
    overdue_wos.sort(key=lambda w: w.days_overdue, reverse=True)

    # Blocked cars — waiting on something
    blocked_statuses = ['Waiting Approval', 'Waiting Material', 'On Hold']
    blocked_wos = [w for w in active_wos if w.status in blocked_statuses]

    # Recent activity
    recent_logs = ActivityLog.query.join(WorkOrder).filter(
        WorkOrder.location_id.in_(loc_filter)
    ).order_by(ActivityLog.created_at.desc()).limit(10).all()

    locations = Location.query.all()

    # ── DATE RANGES ──────────────────────────────────────
    today = date.today()
    month_start = date(today.year, today.month, 1)

    # ── WEEKLY FLOW — last 8 weeks ────────────────────────
    weeks_data = []
    for i in range(7, -1, -1):
        wk_start = today - timedelta(weeks=i, days=today.weekday())
        wk_end   = wk_start + timedelta(days=6)
        arrived  = WorkOrder.query.filter(
            WorkOrder.location_id.in_(loc_filter),
            WorkOrder.date_arrived >= wk_start,
            WorkOrder.date_arrived <= wk_end,
        ).count()
        departed = WorkOrder.query.filter(
            WorkOrder.location_id.in_(loc_filter),
            WorkOrder.billed_out >= wk_start,
            WorkOrder.billed_out <= wk_end,
        ).count()
        weeks_data.append({
            'label':    wk_start.strftime('%d %b'),
            'arrived':  arrived,
            'departed': departed,
        })

    # ── THIS MONTH STATS ──────────────────────────────────
    month_completed = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.billed_out >= month_start,
        WorkOrder.billed_out <= today,
    ).count()

    month_revenue = db.session.query(
        db.func.sum(WorkOrder.estimated_revenue)
    ).filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.billed_out >= month_start,
    ).scalar() or 0

    month_arrived = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.date_arrived >= month_start,
    ).count()

    primary_loc_id = loc_filter[0] if loc_filter else None
    monthly_target = MonthlyTarget.query.filter_by(
        location_id=primary_loc_id,
        month=today.month,
        year=today.year,
    ).first()
    car_target     = monthly_target.car_target     if monthly_target else 40
    revenue_target = monthly_target.revenue_target if monthly_target else 500000

    # ── PIPELINE BOTTLENECK ────────────────────────────────
    pipeline_stages = [
        'To Arrive', 'Arrived', 'Waiting Inspection',
        'Waiting Estimate', 'Waiting Approval',
        'Waiting Work Order', 'Work Order Created',
        'Waiting Cleaning', 'Cleaned', 'Waiting Repair',
        'Waiting Material', 'In Repair', 'Repair Complete',
        'Waiting Dispo', 'Dispo Received', 'Ready to Bill',
    ]
    pipeline_data = []
    for stage in pipeline_stages:
        count = WorkOrder.query.filter(
            WorkOrder.location_id.in_(loc_filter),
            WorkOrder.status == stage,
            WorkOrder.is_closed == False,
        ).count()
        if count > 0:
            pipeline_data.append({'stage': stage, 'count': count})

    # ── PURPOSE BREAKDOWN ─────────────────────────────────
    purposes = [
        'Bad Order', 'Full Qualification', 'Partial Qualification',
        'Railroad Damage', 'Lining Inspection', 'Reline', 'Cleaning',
    ]
    purpose_data = []
    for p in purposes:
        count = WorkOrder.query.filter(
            WorkOrder.location_id.in_(loc_filter),
            WorkOrder.purpose == p,
            WorkOrder.is_closed == False,
        ).count()
        purpose_data.append({'purpose': p, 'count': count})

    # ── LOCATION COMPARISON ───────────────────────────────
    location_stats = []
    for loc in Location.query.all():
        active_count = WorkOrder.query.filter_by(
            location_id=loc.id, is_closed=False
        ).count()
        completed_m = WorkOrder.query.filter(
            WorkOrder.location_id == loc.id,
            WorkOrder.billed_out >= month_start,
        ).count()
        spots_total = Spot.query.filter_by(
            location_id=loc.id, is_active=True
        ).count()
        occupied = SpotAssignment.query.join(Spot).filter(
            Spot.location_id == loc.id,
            SpotAssignment.is_active == True,
        ).count()
        utilization = round(occupied / spots_total * 100) if spots_total > 0 else 0
        location_stats.append({
            'name':            loc.name,
            'active':          active_count,
            'completed_month': completed_m,
            'utilization':     utilization,
            'occupied':        occupied,
            'total_spots':     spots_total,
        })

    # ── AGING BREAKDOWN ───────────────────────────────────
    all_active_wos = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.is_closed == False,
        WorkOrder.date_arrived != None,
    ).all()
    aging_data = {
        'green': sum(1 for w in all_active_wos if (w.days_on_site or 0) < 30),
        'amber': sum(1 for w in all_active_wos if 30 <= (w.days_on_site or 0) < 60),
        'red':   sum(1 for w in all_active_wos if (w.days_on_site or 0) >= 60),
    }

    # ── TURNAROUND TIMES (last ~3 months of completed WOs) ─
    if today.month > 2:
        three_months_ago = date(today.year, today.month - 2, 1)
    else:
        three_months_ago = date(today.year, 1, 1)

    completed_wos = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.billed_out != None,
        WorkOrder.date_arrived != None,
        WorkOrder.billed_out >= three_months_ago,
    ).all()

    if completed_wos:
        avg_turnaround = round(
            sum((w.billed_out - w.date_arrived).days for w in completed_wos)
            / len(completed_wos)
        )
        wo_cs = [w for w in completed_wos if w.work_order_created]
        avg_cs_processing = round(
            sum((w.work_order_created - w.date_arrived).days for w in wo_cs)
            / len(wo_cs)
        ) if wo_cs else 0
        wo_sched = [w for w in completed_wos if w.repair_start_date and w.work_order_created]
        avg_sched_lag = round(
            sum((w.repair_start_date - w.work_order_created).days for w in wo_sched)
            / len(wo_sched)
        ) if wo_sched else 0
    else:
        avg_turnaround    = 0
        avg_cs_processing = 0
        avg_sched_lag     = 0

    # ── PARTS BLOCKED ─────────────────────────────────────
    parts_blocked = WorkOrder.query.filter(
        WorkOrder.location_id.in_(loc_filter),
        WorkOrder.parts_status.in_(['Ordered', 'Waiting ETA', 'Partially Available']),
        WorkOrder.is_closed == False,
    ).count()

    # ── TOTAL IN YARD ─────────────────────────────────────
    total_in_yard = YardCar.query.join(WorkOrder).filter(
        WorkOrder.location_id.in_(loc_filter),
        YardCar.is_active == True,
    ).count()

    # ── TOP CUSTOMERS ─────────────────────────────────────
    customer_counts  = defaultdict(int)
    customer_revenue = defaultdict(float)
    for wo in active_wos:
        if wo.car and wo.car.customer:
            name = wo.car.customer.name
            customer_counts[name]  += 1
            customer_revenue[name] += wo.estimated_revenue or 0

    top_customers = sorted(
        [{'name': k, 'count': customer_counts[k], 'revenue': customer_revenue[k]}
         for k in customer_counts],
        key=lambda x: x['revenue'], reverse=True
    )[:6]

    return render_template('dashboard.html',
        user=user,
        stats=stats,
        overdue_wos=overdue_wos,
        blocked_wos=blocked_wos,
        recent_logs=recent_logs,
        active_wos=active_wos,
        locations=locations,
        get_status_color=get_status_color,
        # Chart / analytics data
        weeks_data=weeks_data,
        pipeline_data=pipeline_data,
        purpose_data=purpose_data,
        location_stats=location_stats,
        aging_data=aging_data,
        avg_turnaround=avg_turnaround,
        avg_cs_processing=avg_cs_processing,
        avg_sched_lag=avg_sched_lag,
        parts_blocked=parts_blocked,
        total_in_yard=total_in_yard,
        top_customers=top_customers,
        month_completed=month_completed,
        month_revenue=month_revenue,
        month_arrived=month_arrived,
        car_target=car_target,
        revenue_target=revenue_target,
    )


# ─────────────────────────────────────────────
# CAR TRACKING — list all work orders
# ─────────────────────────────────────────────

@main.route('/car-tracking')
def car_tracking():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id   = request.args.get('location_id', type=int)
    status_filter = request.args.get('status', '')
    search        = request.args.get('search', '').strip()

    query = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == False)

    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    if status_filter:
        query = query.filter(WorkOrder.status == status_filter)
    if search:
        query = query.filter(db.or_(
            Car.car_mark.ilike(f'%{search}%'),
            Car.car_number.ilike(f'%{search}%'),
            db.cast(Car.car_number, db.String).ilike(f'%{search}%'),
        ))

    all_wos = query.order_by(WorkOrder.date_arrived.desc()).all()

    # Recently closed — last 30 days for reference
    recent_closed = WorkOrder.query.join(Car).filter(
        WorkOrder.is_closed == True,
        WorkOrder.billed_out >= date.today() - timedelta(days=30),
    ).order_by(WorkOrder.billed_out.desc()).all()

    # ── CSR ACTION QUEUE ──────────────────────────────────
    def days_since(d):
        return (date.today() - d).days if d else None

    action_queue = {
        'request_pkg':          [],
        'create_estimate':      [],
        'follow_up_estimate':   [],
        'create_work_order':    [],
        'request_dispo':        [],
        'send_final_bill':      [],
        'await_final_approval': [],
    }

    for wo in all_wos:
        if wo.date_arrived and not wo.shopping_pkg_received:
            action_queue['request_pkg'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.date_arrived),
                'action_label': 'Request Shopping Package',
                'stamp_field': 'shopping_pkg_received',
                'color': 'blue',
            })
        elif wo.inbound_inspection and not wo.estimate_sent:
            action_queue['create_estimate'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.inbound_inspection),
                'action_label': 'Send Estimate',
                'stamp_field': 'estimate_sent',
                'color': 'yellow',
            })
        elif (wo.estimate_sent and not wo.estimate_approved
              and (days_since(wo.estimate_sent) or 0) >= 7):
            action_queue['follow_up_estimate'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.estimate_sent),
                'action_label': 'Follow Up with Customer',
                'stamp_field': None,
                'color': 'orange',
            })
        elif wo.estimate_approved and not wo.work_order_created:
            action_queue['create_work_order'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.estimate_approved),
                'action_label': 'Create Work Order',
                'stamp_field': 'work_order_created',
                'color': 'teal',
            })
        elif wo.outbound_inspection and not wo.dispo_requested:
            action_queue['request_dispo'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.outbound_inspection),
                'action_label': 'Request Dispo',
                'stamp_field': 'dispo_requested',
                'color': 'green',
            })
        elif wo.dispo_received and not wo.final_estimate_sent:
            action_queue['send_final_bill'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.dispo_received),
                'action_label': 'Send Final Bill',
                'stamp_field': 'final_estimate_sent',
                'color': 'green',
            })
        elif wo.final_estimate_sent and not wo.final_estimate_approved:
            action_queue['await_final_approval'].append({
                'wo': wo, 'car': wo.car,
                'days_waiting': days_since(wo.final_estimate_sent),
                'action_label': 'Awaiting Final Approval',
                'stamp_field': 'final_estimate_approved',
                'color': 'purple',
            })

    total_actions = sum(len(v) for v in action_queue.values())

    locations = Location.query.all()
    all_statuses = [
        'To Arrive', 'Arrived', 'Waiting Inspection', 'Waiting Estimate',
        'Waiting Approval', 'Waiting Work Order', 'Work Order Created',
        'Waiting Cleaning', 'Cleaned', 'Waiting Repair', 'Waiting Material',
        'In Repair', 'Repair Complete', 'Waiting Outbound Inspection',
        'Waiting Dispo', 'Dispo Received', 'Ready to Bill', 'Billed',
        'Completed', 'On Hold', 'Cancelled',
    ]

    return render_template('car_tracking.html',
        user=user,
        all_wos=all_wos,
        recent_closed=recent_closed,
        action_queue=action_queue,
        total_actions=total_actions,
        locations=locations,
        all_statuses=all_statuses,
        current_location=location_id,
        current_status=status_filter,
        search=search,
        get_status_color=get_status_color,
    )


# ─────────────────────────────────────────────
# CSR DATE STAMP
# ─────────────────────────────────────────────

@main.route('/car-tracking/stamp/<int:wo_id>', methods=['POST'])
def stamp_date(wo_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    wo    = WorkOrder.query.get_or_404(wo_id)
    field = request.form.get('field')

    allowed_fields = [
        'shopping_pkg_received', 'estimate_sent', 'estimate_approved',
        'work_order_created', 'inbound_inspection', 'dispo_requested',
        'dispo_received', 'final_estimate_sent', 'final_estimate_approved',
        'outbound_inspection', 'billed_out',
    ]

    if field not in allowed_fields:
        flash('Invalid field.', 'warning')
        return redirect(url_for('main.car_tracking'))

    old_val = getattr(wo, field)
    new_val = date.today()
    setattr(wo, field, new_val)
    wo.updated_by = user.name

    log_activity(wo.id, user.name, user.role,
                 f'Stamped {field} via action queue',
                 field, old_val, new_val)

    auto_calculate_status(wo)
    db.session.commit()

    flash(
        f'{field.replace("_", " ").title()} marked as today '
        f'({new_val.strftime("%b %d, %Y")}). '
        f'Edit on car profile if date is incorrect.',
        'success'
    )
    return redirect(url_for('main.car_tracking'))


# ─────────────────────────────────────────────
# CAR PROFILE — full detail for one work order
# ─────────────────────────────────────────────

@main.route('/car/<int:wo_id>')
def car_profile(wo_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    wo  = WorkOrder.query.get_or_404(wo_id)
    car = wo.car

    # All visits for this physical car
    all_visits = WorkOrder.query.filter_by(car_id=car.id)\
                    .order_by(WorkOrder.visit_number).all()

    comments    = Comment.query.filter_by(work_order_id=wo.id)\
                    .order_by(Comment.created_at.desc()).all()
    activity    = ActivityLog.query.filter_by(work_order_id=wo.id)\
                    .order_by(ActivityLog.created_at.desc()).all()
    spot_assign = SpotAssignment.query.filter_by(work_order_id=wo.id)\
                    .order_by(SpotAssignment.assigned_at.desc()).first()

    return render_template('car_profile.html',
        user=user,
        wo=wo,
        car=car,
        all_visits=all_visits,
        comments=comments,
        activity=activity,
        spot_assign=spot_assign,
        get_status_color=get_status_color,
    )


# ─────────────────────────────────────────────
# ADD NEW CAR
# ─────────────────────────────────────────────

@main.route('/car/new', methods=['GET', 'POST'])
def add_car():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    if request.method == 'POST':
        mark   = request.form.get('car_mark', '').strip().upper()
        number = request.form.get('car_number', '').strip()

        # ── Server-side validation ────────────────────────────────
        errors = []
        if not mark:
            errors.append('Car mark is required.')
        elif not mark.isalpha():
            errors.append('Car mark must contain only letters.')
        if not number:
            errors.append('Car number is required.')
        elif not number.isdigit():
            errors.append('Car number must contain only digits.')

        rev_raw  = request.form.get('estimated_revenue', '').strip()
        hrs_raw  = request.form.get('estimated_hours', '').strip()
        if rev_raw:
            try:
                float(rev_raw)
            except ValueError:
                errors.append('Estimated revenue must be a number.')
        if hrs_raw:
            try:
                float(hrs_raw)
            except ValueError:
                errors.append('Estimated hours must be a number.')

        date_arrived_raw = request.form.get('date_arrived', '').strip()
        if date_arrived_raw:
            try:
                date.fromisoformat(date_arrived_raw)
            except ValueError:
                errors.append('Arrival date must be a valid date (YYYY-MM-DD).')

        if errors:
            for e in errors:
                flash(e, 'error')
            return redirect(url_for('main.add_car'))

        # Check if car already exists
        existing_car = Car.query.filter_by(car_mark=mark, car_number=number).first()

        if existing_car:
            # Car exists — check if it has an open work order
            open_wo = existing_car.active_work_order
            if open_wo:
                flash(f'{mark}-{number} already has an open work order (WO #{open_wo.id}). '
                      f'Update that one instead.', 'warning')
                return redirect(url_for('main.car_profile', wo_id=open_wo.id))
            # Car exists but no open WO — create a new visit
            car         = existing_car
            visit_num   = len(existing_car.work_orders) + 1
            is_returning = True
        else:
            # Brand new car
            loc_id = request.form.get('location_id', type=int) or user.location_id
            cust_name = request.form.get('customer_name', '').strip()
            customer = Customer.query.filter_by(name=cust_name).first()
            if not customer and cust_name:
                customer = Customer(name=cust_name)
                db.session.add(customer)
                db.session.flush()

            car = Car(
                location_id=loc_id,
                customer_id=customer.id if customer else None,
                car_mark=mark,
                car_number=number,
                commodity=request.form.get('commodity', '').strip(),
                created_by=user.name,
            )
            db.session.add(car)
            db.session.flush()
            visit_num    = 1
            is_returning = False

        # Create the work order
        loc_id = request.form.get('location_id', type=int) or user.location_id
        wo = WorkOrder(
            car_id=car.id,
            location_id=loc_id,
            visit_number=visit_num,
            reason_shopped=request.form.get('reason_shopped', '').strip(),
            scope_of_work=request.form.get('scope_of_work', '').strip(),
            purpose=request.form.get('purpose', ''),
            priority=int(request.form.get('priority', 3)),
            estimated_revenue=float(request.form.get('estimated_revenue') or 0),
            estimated_hours=float(request.form.get('estimated_hours') or 0),
            status='Inbounded',
            date_arrived=_parse_date(request.form.get('date_arrived')),
            shopping_pkg_received=_parse_date(request.form.get('shopping_pkg_received')),
            created_by=user.name,
        )
        db.session.add(wo)
        db.session.flush()

        log_activity(wo.id, user.name, user.role,
                     f'Created work order (Visit #{visit_num})')

        db.session.commit()

        msg = f'Returning car — Visit #{visit_num} created.' if is_returning else 'Car added successfully.'
        flash(msg, 'success')
        return redirect(url_for('main.car_profile', wo_id=wo.id))

    locations = Location.query.all()
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('add_car.html',
        user=user,
        locations=locations,
        customers=customers,
    )


# ─────────────────────────────────────────────
# UPDATE WORK ORDER DATES / STATUS
# ─────────────────────────────────────────────

@main.route('/car/<int:wo_id>/update', methods=['POST'])
def update_work_order(wo_id):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Not logged in'}), 401

    wo = WorkOrder.query.get_or_404(wo_id)
    was_closed = wo.is_closed  # Track for post-close correction log

    # Date fields that can be updated
    date_fields = [
        'date_arrived', 'shopping_pkg_received', 'inbound_inspection',
        'estimate_sent', 'estimate_approved', 'work_order_created',
        'cleaning_scheduled', 'cleaning_completed',
        'repair_start_date', 'repair_completed_date', 'projected_outbound',
        'outbound_inspection', 'dispo_requested', 'dispo_received',
        'billed_out', 'final_estimate_sent', 'final_estimate_approved',
    ]

    updated_fields = []

    for field in date_fields:
        if field in request.form:
            new_val  = _parse_date(request.form.get(field))
            old_val  = getattr(wo, field)
            if new_val != old_val:
                log_activity(wo.id, user.name, user.role,
                             f'Updated {field}', field, old_val, new_val)
                setattr(wo, field, new_val)
                updated_fields.append(field)

    # Status update
    if 'status' in request.form:
        new_status = request.form.get('status')
        if new_status != wo.status:
            log_activity(wo.id, user.name, user.role,
                         'Updated status', 'status', wo.status, new_status)
            wo.status = new_status
            if new_status in ('Completed', 'Cancelled'):
                wo.is_closed = True

    # Other text fields
    for field in ['scope_of_work', 'reason_shopped', 'purpose',
                  'estimated_revenue', 'estimated_hours', 'actual_hours',
                  'car_type', 'parts_status', 'parts_notes',
                  'rework_notes', 'rework_date', 'parts_eta',
                  'bp_blast_hrs', 'bp_interior_coating_hrs', 'bp_lining_touchup_hrs',
                  'bp_buff_lining_hrs', 'bp_paint_booth_hrs', 'bp_notes']:
        if field in request.form:
            new_val = request.form.get(field)
            old_val = str(getattr(wo, field) or '')
            if new_val != old_val:
                log_activity(wo.id, user.name, user.role,
                             f'Updated {field}', field, old_val, new_val)
                if field in ('estimated_revenue', 'estimated_hours', 'actual_hours',
                             'bp_blast_hrs', 'bp_interior_coating_hrs', 'bp_lining_touchup_hrs',
                             'bp_buff_lining_hrs', 'bp_paint_booth_hrs'):
                    setattr(wo, field, float(new_val) if new_val else 0.0)
                elif field in ('rework_date', 'parts_eta'):
                    setattr(wo, field, _parse_date(new_val))
                else:
                    setattr(wo, field, new_val)

    # Checkbox fields — each form section uses _section to declare which checkboxes it owns.
    # Absent checkboxes within a submitted section = unchecked (False).
    section = request.form.get('_section', '')
    section_cb_map = {
        'main':   ['cleaning_required'],
        'rework': ['rework_required', 'rework_passed'],
        'bp':     ['blast_paint_required'],
    }
    cb_fields = section_cb_map.get(section, [])
    for cb_field in cb_fields:
        new_bool = request.form.get(cb_field) == 'on'
        if new_bool != getattr(wo, cb_field):
            log_activity(wo.id, user.name, user.role,
                         f'Updated {cb_field}', cb_field,
                         getattr(wo, cb_field), new_bool)
            setattr(wo, cb_field, new_bool)

    # Don't auto-calculate over manually managed statuses
    manual_statuses = {'On Hold', 'Cancelled', 'Waiting Repair', 'Waiting Material', 'Waiting Approval'}
    if wo.status not in manual_statuses:
        auto_calculate_status(wo)

    # Log post-close corrections so there's an audit trail
    if was_closed and updated_fields:
        log_activity(wo.id, user.name, user.role,
                     f'Post-close correction — fields updated: {", ".join(updated_fields)}')
        wo.is_closed = False if not wo.billed_out else True  # Reopen only if billed_out was cleared

    wo.updated_by = user.name
    wo.updated_at = datetime.utcnow()
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': True,
            'status': wo.status,
            'status_color': get_status_color(wo.status),
            'completion': wo.completion_percentage,
            'updated_fields': updated_fields,
        })

    flash('Work order updated.', 'success')
    return redirect(url_for('main.car_profile', wo_id=wo_id))


# ─────────────────────────────────────────────
# ADD COMMENT
# ─────────────────────────────────────────────

@main.route('/car/<int:wo_id>/comment', methods=['POST'])
def add_comment(wo_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    body = request.form.get('body', '').strip()
    if body:
        comment = Comment(
            work_order_id=wo_id,
            author=user.name,
            role=user.role,
            body=body,
        )
        db.session.add(comment)
        log_activity(wo_id, user.name, user.role, 'Added comment')
        db.session.commit()

    return redirect(url_for('main.car_profile', wo_id=wo_id))


# ─────────────────────────────────────────────
# SHOP FLOOR
# ─────────────────────────────────────────────

@main.route('/shop-floor')
def shop_floor():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id

    location  = Location.query.get(location_id)
    spots     = Spot.query.filter_by(location_id=location_id, is_active=True).all()
    locations = Location.query.all()

    # Build zone-grouped spot data
    zones = OrderedDict()
    assigned_wo_ids = set()

    for spot in spots:
        zone = get_spot_zone(spot.name, location.name)
        if zone not in zones:
            zones[zone] = {'spots': [], 'occupied': 0, 'total': 0}

        assignment = spot.current_assignment
        wo  = WorkOrder.query.get(assignment.work_order_id) if assignment else None
        car = wo.car if wo else None

        # Auto-progress from repair_start_date → target_date
        auto_pct = None
        if assignment and wo and wo.repair_start_date and assignment.target_date:
            total_days = (assignment.target_date - wo.repair_start_date).days
            elapsed    = (date.today() - wo.repair_start_date).days
            if total_days > 0:
                auto_pct = min(100, round(elapsed / total_days * 100))

        zones[zone]['spots'].append({
            'spot':       spot,
            'assignment': assignment,
            'wo':         wo,
            'car':        car,
            'auto_pct':   auto_pct,
            'warnings':   get_scheduling_warnings(wo) if wo else [],
        })
        zones[zone]['total'] += 1
        if assignment:
            zones[zone]['occupied'] += 1
            assigned_wo_ids.add(assignment.work_order_id)

    # Next car queue from schedule entries not yet in a spot
    next_queue_entries = ScheduleEntry.query.filter(
        ScheduleEntry.location_id == location_id,
        ScheduleEntry.schedule_status.notin_(['Completed', 'Pushed to Next Month'])
    ).order_by(
        ScheduleEntry.target_week,
        ScheduleEntry.rev_per_hour.desc()
    ).limit(20).all()

    next_queue = []
    for e in next_queue_entries:
        if e.work_order_id not in assigned_wo_ids:
            wo = WorkOrder.query.get(e.work_order_id)
            if wo and not wo.is_closed:
                next_queue.append({
                    'entry':    e,
                    'wo':       wo,
                    'car':      wo.car,
                    'warnings': get_scheduling_warnings(wo),
                })
        if len(next_queue) >= 10:
            break

    # Available WOs for manual assignment — any active non-completed car
    available_wos = WorkOrder.query.filter(
        WorkOrder.location_id == location_id,
        WorkOrder.is_closed == False,
        WorkOrder.status.notin_(['Completed', 'Cancelled', 'Billed']),
    ).all()
    available_wos = [w for w in available_wos if w.id not in assigned_wo_ids]

    total_spots    = sum(z['total']    for z in zones.values())
    occupied_spots = sum(z['occupied'] for z in zones.values())

    # ── Cleaning rack slot map: spot_id → CleaningEntry (for FIX 4D sync) ──
    cleaning_entries_active = CleaningEntry.query.filter(
        CleaningEntry.location_id == location_id,
        CleaningEntry.cleaning_status.notin_(['Cancelled']),
    ).all()
    cleaning_pos_map = {}
    for ce in cleaning_entries_active:
        if ce.rack_position:
            cleaning_pos_map[ce.rack_position] = ce

    cleaning_slot_map = {}
    for spot in spots:
        if spot.name.startswith('CL-'):
            try:
                rack_pos = int(spot.name.split('-')[1])
                cleaning_slot_map[spot.id] = cleaning_pos_map.get(rack_pos)
            except (IndexError, ValueError):
                pass

    return render_template('shop_floor.html',
        user=user,
        location=location,
        locations=locations,
        zones=zones,
        next_queue=next_queue,
        available_wos=available_wos,
        total_spots=total_spots,
        occupied_spots=occupied_spots,
        cleaning_slot_map=cleaning_slot_map,
        get_status_color=get_status_color,
        get_scheduling_warnings=get_scheduling_warnings,
    )


# ─────────────────────────────────────────────
# ASSIGN CAR TO SPOT
# ─────────────────────────────────────────────

@main.route('/shop-floor/assign', methods=['POST'])
def assign_spot():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    spot_id      = request.form.get('spot_id', type=int)
    wo_id        = request.form.get('wo_id', type=int)
    location_id  = request.form.get('location_id', type=int)

    if not spot_id:
        flash('No spot selected. Please select a spot before assigning.', 'warning')
        return redirect(url_for('main.shop_floor', location_id=location_id))

    if not wo_id:
        flash('No car selected. Please select a car before assigning.', 'warning')
        return redirect(url_for('main.shop_floor', location_id=location_id))

    # Verify spot exists
    spot = Spot.query.get(spot_id)
    if not spot:
        flash(f'Spot ID {spot_id} not found.', 'warning')
        return redirect(url_for('main.shop_floor', location_id=location_id))

    # Verify work order exists
    wo = WorkOrder.query.get(wo_id)
    if not wo:
        flash('Work order not found.', 'warning')
        return redirect(url_for('main.shop_floor', location_id=location_id))

    sow_type     = request.form.get('sow_type', '')
    technician   = request.form.get('technician', '').strip()
    est_hours    = float(request.form.get('estimated_hours') or 0)
    target_date  = _parse_date(request.form.get('target_date'))
    notes        = request.form.get('notes', '').strip()

    # Clear any existing active assignment for this spot
    existing = SpotAssignment.query.filter_by(spot_id=spot_id, is_active=True).first()
    if existing:
        existing.is_active  = False
        existing.cleared_at = datetime.utcnow()

    assignment = SpotAssignment(
        spot_id=spot_id,
        work_order_id=wo_id,
        sow_type=sow_type,
        technician=technician,
        estimated_hours=est_hours,
        target_date=target_date,
        notes=notes,
        status='Under Repair',
        is_active=True,
        assigned_by=user.name,
    )
    db.session.add(assignment)

    # Update work order status (wo already fetched above)
    if wo.status not in ('In Repair', 'Waiting Material'):
        old_status = wo.status
        wo.status  = 'In Repair'
        log_activity(wo_id, user.name, user.role,
                     'Assigned to spot', 'status', old_status, 'In Repair')

    db.session.commit()
    flash('Car assigned to spot.', 'success')
    return redirect(url_for('main.shop_floor', location_id=location_id))


# ─────────────────────────────────────────────
# UPDATE SPOT PROGRESS
# ─────────────────────────────────────────────

@main.route('/shop-floor/update-spot/<int:assignment_id>', methods=['POST'])
def update_spot(assignment_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    assignment = SpotAssignment.query.get_or_404(assignment_id)

    assignment.pct_complete  = int(request.form.get('pct_complete') or 0)
    assignment.actual_hours  = float(request.form.get('actual_hours') or 0)
    assignment.status        = request.form.get('status', assignment.status)
    assignment.hold_reason   = request.form.get('hold_reason', '').strip()
    assignment.notes         = request.form.get('notes', '').strip()
    assignment.updated_by    = user.name

    # Sync work order status
    wo = WorkOrder.query.get(assignment.work_order_id)
    if wo:
        spot_status = assignment.status
        if spot_status == 'Completed':
            # Set repair date then auto-calculate
            if not wo.repair_completed_date:
                wo.repair_completed_date = date.today()
            log_activity(wo.id, user.name, user.role,
                         'Spot completed — repair date set', 'status', wo.status, 'Repair Complete')
            auto_calculate_status(wo)
            wo.updated_by = user.name
        elif spot_status in ('On Hold', 'Waiting Material'):
            # Manual override statuses — set directly
            override_map = {'On Hold': 'On Hold', 'Waiting Material': 'Waiting Material'}
            new_wo_status = override_map[spot_status]
            if new_wo_status != wo.status:
                log_activity(wo.id, user.name, user.role,
                             'Spot status update', 'status', wo.status, new_wo_status)
                wo.status     = new_wo_status
                wo.updated_by = user.name
        else:
            # All other spot statuses — auto-calculate from dates
            direct_map = {
                'Under Repair':              'In Repair',
                'Waiting Customer Approval': 'Waiting Approval',
                'Waiting Cleaning':          'Waiting Cleaning',
                'Rework Needed':             'In Repair',
            }
            mapped = direct_map.get(spot_status)
            if mapped and mapped != wo.status:
                log_activity(wo.id, user.name, user.role,
                             'Spot status update', 'status', wo.status, mapped)
                wo.status     = mapped
                wo.updated_by = user.name
            auto_calculate_status(wo)

    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'pct': assignment.pct_complete})

    location_id = request.form.get('location_id', type=int)
    return redirect(url_for('main.shop_floor', location_id=location_id))


# ─────────────────────────────────────────────
# CLEAR SPOT
# ─────────────────────────────────────────────

@main.route('/shop-floor/clear-spot/<int:assignment_id>', methods=['POST'])
def clear_spot(assignment_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    assignment             = SpotAssignment.query.get_or_404(assignment_id)
    assignment.is_active   = False
    assignment.cleared_at  = datetime.utcnow()
    assignment.updated_by  = user.name

    log_activity(assignment.work_order_id, user.name, user.role, 'Cleared spot')
    db.session.commit()

    location_id = request.form.get('location_id', type=int)
    flash('Spot cleared.', 'success')
    return redirect(url_for('main.shop_floor', location_id=location_id))


# ─────────────────────────────────────────────
# YARD
# ─────────────────────────────────────────────

@main.route('/yard')
def yard():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id

    location  = Location.query.get(location_id)
    locations = Location.query.all()

    # Default track (track_id is NOT NULL — needed for auto-created entries)
    default_track = YardTrack.query.filter_by(location_id=location_id, is_active=True).first()
    default_track_id = default_track.id if default_track else 1

    # All open work orders at this location
    open_wos = WorkOrder.query.filter_by(location_id=location_id, is_closed=False).all()

    # Find which WO IDs already have an active YardCar
    active_yc_wo_ids = set(
        row[0] for row in
        db.session.query(YardCar.work_order_id)
            .join(WorkOrder, YardCar.work_order_id == WorkOrder.id)
            .filter(WorkOrder.location_id == location_id, YardCar.is_active == True).all()
    )

    # Auto-create YardCar for any open WO that lacks one
    new_entries = False
    for wo in open_wos:
        if wo.id not in active_yc_wo_ids:
            initial_status = 'Arrived' if wo.date_arrived else 'To Arrive'
            yc = YardCar(
                work_order_id=wo.id,
                track_id=default_track_id,
                yard_status=initial_status,
                is_active=True,
                entered_at=datetime.utcnow(),
                updated_by='system',
            )
            db.session.add(yc)
            new_entries = True
    if new_entries:
        db.session.commit()

    # Fetch all active yard cars for this location
    raw = YardCar.query \
              .join(WorkOrder, YardCar.work_order_id == WorkOrder.id) \
              .filter(WorkOrder.location_id == location_id, YardCar.is_active == True) \
              .order_by(YardCar.entered_at.desc()).all()

    yard_cars = []
    for yc in raw:
        wo  = WorkOrder.query.get(yc.work_order_id)
        car = wo.car if wo else None
        yard_cars.append({'yard_car': yc, 'wo': wo, 'car': car})

    # Stats (billed_month comes from inactive YardCars billed this month)
    this_month = date.today().replace(day=1)
    billed_month = YardCar.query \
                       .join(WorkOrder, YardCar.work_order_id == WorkOrder.id) \
                       .filter(WorkOrder.location_id == location_id,
                               YardCar.yard_status == 'Billed',
                               YardCar.exited_at >= this_month).count()
    yard_stats = {
        'total':         len(yard_cars),
        'to_arrive':     sum(1 for y in yard_cars if y['yard_car'].yard_status == 'To Arrive'),
        'ready_to_bill': sum(1 for y in yard_cars if y['yard_car'].yard_status == 'Ready to Bill'),
        'billed_month':  billed_month,
    }

    wo_statuses = [
        'To Arrive', 'Arrived', 'Waiting Inspection', 'Waiting Estimate',
        'Waiting Approval', 'Waiting Work Order', 'Work Order Created',
        'Waiting Cleaning', 'Cleaned', 'Waiting Repair', 'Waiting Material',
        'In Repair', 'Repair Complete', 'Waiting Outbound Inspection',
        'Waiting Dispo', 'Dispo Received', 'Ready to Bill',
        'On Hold', 'Cancelled',
    ]

    return render_template('yard.html',
        user=user,
        location=location,
        locations=locations,
        yard_cars=yard_cars,
        yard_stats=yard_stats,
        wo_statuses=wo_statuses,
        get_status_color=get_status_color,
    )


# ─────────────────────────────────────────────
# ADD CAR TO YARD
# ─────────────────────────────────────────────

@main.route('/yard/add', methods=['POST'])
def add_to_yard():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.form.get('location_id', type=int)

    # track_id is NOT NULL — use first available track for this location as a default
    default_track = YardTrack.query.filter_by(location_id=location_id, is_active=True).first()
    track_id = default_track.id if default_track else 1

    yc = YardCar(
        work_order_id=request.form.get('wo_id', type=int),
        track_id=track_id,
        yard_status=request.form.get('yard_status', 'To Arrive'),
        sow_notes=request.form.get('sow_notes', '').strip(),
        comments=request.form.get('comments', '').strip(),
        return_location=request.form.get('return_location', '').strip(),
        updated_by=user.name,
    )
    db.session.add(yc)
    db.session.commit()

    flash('Car added to yard.', 'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# UPDATE YARD CAR (edit modal — notes/return location)
# ─────────────────────────────────────────────

@main.route('/yard/update/<int:yc_id>', methods=['POST'])
def update_yard_car(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc = YardCar.query.get_or_404(yc_id)
    yc.sow_notes       = request.form.get('sow_notes', yc.sow_notes or '').strip()
    yc.comments        = request.form.get('comments', '').strip()
    yc.return_location = request.form.get('return_location', yc.return_location or '').strip()
    yc.updated_by      = user.name

    db.session.commit()

    location_id = request.form.get('location_id', type=int)
    flash('Yard car updated.', 'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# UPDATE YARD / WO STATUS (inline with validation)
# ─────────────────────────────────────────────

@main.route('/yard/update-status/<int:yc_id>', methods=['POST'])
def yard_update_status(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc  = YardCar.query.get_or_404(yc_id)
    wo  = WorkOrder.query.get(yc.work_order_id)
    location_id = request.form.get('location_id', type=int)

    new_yard_status = request.form.get('yard_status', '').strip()
    new_wo_status   = request.form.get('wo_status', '').strip()

    if new_yard_status and new_yard_status != yc.yard_status:
        if new_yard_status in ('In Yard', 'In Yard Repair') and (not wo or not wo.date_arrived):
            flash('Cannot mark as In Yard — no arrival date on record. '
                  'Add the arrival date on the car profile first.', 'warning')
            return redirect(url_for('main.yard', location_id=location_id))
        if new_yard_status == 'Ready to Bill' and (not wo or not wo.outbound_inspection):
            flash('Cannot mark Ready to Bill — outbound inspection not completed. '
                  'Update the car profile first.', 'warning')
            return redirect(url_for('main.yard', location_id=location_id))
        if new_yard_status == 'Billed':
            flash('Use the Bill Out button to bill a car out.', 'warning')
            return redirect(url_for('main.yard', location_id=location_id))
        log_activity(yc.work_order_id, user.name, user.role,
                     'Updated yard status', 'yard_status', yc.yard_status, new_yard_status)
        yc.yard_status = new_yard_status
        yc.updated_by  = user.name
        # Auto-sync WO status from dates when yard status changes
        if wo:
            auto_calculate_status(wo)
            wo.updated_by = user.name

    if new_wo_status and wo and new_wo_status != wo.status:
        if new_wo_status == 'Billed':
            flash('Use the Bill Out button to bill a car out.', 'warning')
            return redirect(url_for('main.yard', location_id=location_id))
        # Manual WO status override — do not auto-calculate over it
        log_activity(wo.id, user.name, user.role,
                     'Updated WO status from yard', 'status', wo.status, new_wo_status)
        wo.status     = new_wo_status
        wo.updated_by = user.name

    db.session.commit()
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# YARD — READY TO BILL
# ─────────────────────────────────────────────

@main.route('/yard/ready-bill/<int:yc_id>', methods=['POST'])
def yard_ready_bill(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc  = YardCar.query.get_or_404(yc_id)
    wo  = WorkOrder.query.get(yc.work_order_id)
    location_id = request.form.get('location_id', type=int)

    if not wo or not wo.outbound_inspection:
        flash('Cannot mark Ready to Bill — outbound inspection not completed. '
              'Update the car profile first.', 'warning')
        return redirect(url_for('main.yard', location_id=location_id))

    yc.yard_status = 'Ready to Bill'
    yc.updated_by  = user.name
    wo.status      = 'Ready to Bill'
    wo.updated_by  = user.name

    db.session.commit()
    flash('Car marked Ready to Bill.', 'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# YARD — BILL OUT (two-step with date)
# ─────────────────────────────────────────────

@main.route('/yard/bill-out/<int:yc_id>', methods=['POST'])
def yard_bill_out(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc = YardCar.query.get_or_404(yc_id)
    location_id = request.form.get('location_id', type=int)

    billed_date = _parse_date(request.form.get('billed_out_date', '').strip())
    if not billed_date:
        flash('Billed out date is required.', 'error')
        return redirect(url_for('main.yard', location_id=location_id))

    return_location = request.form.get('return_location', '').strip()
    if return_location:
        yc.return_location = return_location
    yc.yard_status = 'Billed'
    yc.is_active   = False
    yc.exited_at   = datetime.utcnow()
    yc.updated_by  = user.name

    wo = WorkOrder.query.get(yc.work_order_id)
    car_label = wo.car.full_id if wo and wo.car else 'Car'
    if wo:
        wo.billed_out = billed_date
        wo.is_closed  = True
        wo.updated_by = user.name
        auto_calculate_status(wo)
        log_activity(wo.id, user.name, user.role,
                     f'Car billed out on {billed_date.strftime("%b %d, %Y")}')

    db.session.commit()
    flash(f'{car_label} billed out on {billed_date.strftime("%b %d, %Y")} and removed from yard.',
          'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# YARD — UNDO BILL OUT
# ─────────────────────────────────────────────

@main.route('/yard/undo-bill-out/<int:yc_id>', methods=['POST'])
def undo_bill_out(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc = YardCar.query.get(yc_id)
    if not yc:
        flash('Yard entry not found.', 'error')
        return redirect(url_for('main.yard'))

    wo = WorkOrder.query.get(yc.work_order_id)
    location_id = request.form.get('location_id', type=int) \
                  or (wo.location_id if wo else None)

    yc.is_active   = True
    yc.exited_at   = None
    yc.yard_status = 'In Yard'
    yc.updated_by  = user.name

    if wo:
        wo.billed_out = None
        wo.is_closed  = False
        wo.updated_by = user.name
        auto_calculate_status(wo)
        log_activity(wo.id, user.name, user.role,
                     'Bill out reversed — car returned to yard')

    db.session.commit()
    flash('Car returned to yard successfully.', 'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# REMOVE CAR FROM YARD
# ─────────────────────────────────────────────

@main.route('/yard/remove/<int:yc_id>', methods=['POST'])
def remove_from_yard(yc_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    yc            = YardCar.query.get_or_404(yc_id)
    yc.is_active  = False
    yc.exited_at  = datetime.utcnow()
    yc.updated_by = user.name
    db.session.commit()

    location_id = request.form.get('location_id', type=int)
    flash('Car removed from yard.', 'success')
    return redirect(url_for('main.yard', location_id=location_id))


# ─────────────────────────────────────────────
# CLEANING RACK
# ─────────────────────────────────────────────

@main.route('/cleaning')
def cleaning():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id

    location  = Location.query.get(location_id)
    locations = Location.query.all()

    # Show all entries except Cancelled — including Cleaned so they stay visible
    entries = CleaningEntry.query.filter(
        CleaningEntry.location_id == location_id,
        CleaningEntry.cleaning_status != 'Cancelled',
    ).order_by(CleaningEntry.scheduled_date).all()

    # Rack capacity per location
    rack_capacity = {'Thornfield': 7, 'Harwick': 5, 'Delverton': 14}
    capacity = rack_capacity.get(location.name, 10)

    # Build rack slot map — only non-cleaned entries fill slots visually
    slot_map = {}
    for entry in entries:
        if entry.rack_position and entry.cleaning_status not in ('Cancelled', 'Cleaned'):
            slot_map[entry.rack_position] = entry

    cleaning_stats = {
        'total_slots':  capacity,
        'occupied':     len([e for e in entries if e.cleaning_status not in ('Cleaned', 'Cancelled')]),
        'scheduled':    sum(1 for e in entries if e.cleaning_status == 'Scheduled'),
        'in_cleaning':  sum(1 for e in entries if e.cleaning_status == 'In Cleaning'),
        'cleaned':      sum(1 for e in entries if e.cleaning_status == 'Cleaned'),
        'ready':        sum(1 for e in entries if e.cleaning_status == 'Ready for Cleaning'),
    }

    # Active WOs available for scheduling — exclude cars already scheduled
    already_scheduled_wo_ids = {e.work_order_id for e in entries}
    active_wos = WorkOrder.query.filter(
        WorkOrder.location_id == location_id,
        WorkOrder.is_closed == False,
        WorkOrder.status.notin_(['Completed', 'Cancelled', 'Billed']),
    ).all()
    active_wos = [w for w in active_wos if w.id not in already_scheduled_wo_ids]

    return render_template('cleaning.html',
        user=user,
        location=location,
        locations=locations,
        entries=entries,
        slot_map=slot_map,
        capacity=capacity,
        cleaning_stats=cleaning_stats,
        active_wos=active_wos,
        get_status_color=get_status_color,
    )


# ─────────────────────────────────────────────
# SCHEDULE CAR FOR CLEANING
# ─────────────────────────────────────────────

@main.route('/cleaning/add', methods=['POST'])
def add_cleaning():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.form.get('location_id', type=int)
    wo_id       = request.form.get('wo_id', type=int)

    # Prevent duplicate — check for any non-cancelled entry for this WO
    existing = CleaningEntry.query.filter_by(work_order_id=wo_id).filter(
        CleaningEntry.cleaning_status != 'Cancelled'
    ).first()
    if existing:
        flash('This car is already in the cleaning schedule. '
              'Update the existing entry instead.', 'warning')
        return redirect(url_for('main.cleaning', location_id=location_id))

    # Check slot isn't already occupied
    requested_pos = request.form.get('rack_position', type=int)
    if requested_pos:
        slot_taken = CleaningEntry.query.filter_by(
            location_id=location_id, rack_position=requested_pos
        ).filter(CleaningEntry.cleaning_status != 'Cancelled').first()
        if slot_taken:
            taken_wo  = WorkOrder.query.get(slot_taken.work_order_id)
            taken_car = taken_wo.car.full_id if taken_wo and taken_wo.car else '?'
            flash(f'Slot {requested_pos} is already occupied by {taken_car}. '
                  f'Clear that slot first.', 'warning')
            return redirect(url_for('main.cleaning', location_id=location_id))

    entry = CleaningEntry(
        work_order_id=wo_id,
        location_id=location_id,
        rack_position=requested_pos,
        scheduled_date=_parse_date(request.form.get('scheduled_date')),
        cleaning_status='Scheduled',
        material_notes=request.form.get('material_notes', '').strip(),
        notes=request.form.get('notes', '').strip(),
        updated_by=user.name,
    )
    db.session.add(entry)

    wo = WorkOrder.query.get(wo_id)
    if wo:
        wo.cleaning_scheduled = entry.scheduled_date
        auto_calculate_status(wo)
        wo.updated_by = user.name

    db.session.commit()
    flash('Car scheduled for cleaning.', 'success')
    return redirect(url_for('main.cleaning', location_id=location_id))


# ─────────────────────────────────────────────
# UPDATE CLEANING STATUS
# ─────────────────────────────────────────────

@main.route('/cleaning/update/<int:entry_id>', methods=['POST'])
def update_cleaning(entry_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    entry    = CleaningEntry.query.get_or_404(entry_id)
    location_id = request.form.get('location_id', type=int) or entry.location_id

    # Handle rack_position assignment — check for conflicts
    new_pos = request.form.get('rack_position', type=int)
    if new_pos and new_pos != entry.rack_position:
        conflict = CleaningEntry.query.filter(
            CleaningEntry.location_id == entry.location_id,
            CleaningEntry.rack_position == new_pos,
            CleaningEntry.id != entry.id,
            CleaningEntry.cleaning_status != 'Cancelled',
        ).first()
        if conflict:
            conflict_wo  = WorkOrder.query.get(conflict.work_order_id)
            conflict_car = conflict_wo.car.full_id if conflict_wo and conflict_wo.car else '?'
            flash(f'Slot {new_pos} is already occupied by {conflict_car}. '
                  f'Clear that slot first.', 'warning')
            return redirect(url_for('main.cleaning', location_id=location_id))
        entry.rack_position = new_pos
    elif request.form.get('rack_position') == '' and entry.rack_position:
        # Clearing the slot position
        entry.rack_position = None

    entry.cleaning_status  = request.form.get('cleaning_status', entry.cleaning_status)
    entry.notes            = request.form.get('notes', '').strip()
    entry.updated_by       = user.name

    wo = WorkOrder.query.get(entry.work_order_id)
    if entry.cleaning_status == 'Cleaned':
        entry.cleaned_at = datetime.utcnow()
        if wo:
            wo.cleaning_completed = date.today()
            auto_calculate_status(wo)
            wo.updated_by = user.name
    elif wo:
        # Sync cleaning_scheduled from entry date if not already set
        if entry.scheduled_date and not wo.cleaning_scheduled:
            wo.cleaning_scheduled = entry.scheduled_date
            auto_calculate_status(wo)
            wo.updated_by = user.name

    db.session.commit()

    location_id = request.form.get('location_id', type=int) or entry.location_id
    return redirect(url_for('main.cleaning', location_id=location_id))


# ─────────────────────────────────────────────
# REMOVE FROM CLEANING SCHEDULE
# ─────────────────────────────────────────────

@main.route('/cleaning/remove/<int:entry_id>', methods=['POST'])
def remove_cleaning(entry_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    entry       = CleaningEntry.query.get_or_404(entry_id)
    location_id = entry.location_id
    wo          = WorkOrder.query.get(entry.work_order_id)

    entry.cleaning_status = 'Cancelled'
    entry.rack_position   = None

    if wo:
        wo.cleaning_scheduled = None
        auto_calculate_status(wo)
        wo.updated_by = user.name
        log_activity(wo.id, user.name, user.role, 'Removed from cleaning schedule')

    db.session.commit()
    flash('Car removed from cleaning schedule.', 'success')
    return redirect(url_for('main.cleaning', location_id=location_id))


# ─────────────────────────────────────────────
# MONTHLY SCHEDULE
# ─────────────────────────────────────────────

@main.route('/schedule')
def schedule():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id

    month = request.args.get('month', type=int, default=date.today().month)
    year  = request.args.get('year',  type=int, default=date.today().year)
    view  = request.args.get('view', default='board')

    location  = Location.query.get(location_id)
    locations = Location.query.all()

    entries = ScheduleEntry.query.filter_by(
        location_id=location_id, month=month, year=year
    ).order_by(ScheduleEntry.target_week, ScheduleEntry.rev_per_hour.desc()).all()

    # Group by week for board view — store dicts with wo/warnings
    weeks = {1: [], 2: [], 3: [], 4: []}
    for entry in entries:
        week = entry.target_week or 1
        if week in weeks:
            wo       = WorkOrder.query.get(entry.work_order_id)
            car      = wo.car if wo else None
            warnings = get_scheduling_warnings(wo) if wo else []
            weeks[week].append({'entry': entry, 'wo': wo, 'car': car, 'warnings': warnings})

    # Unscheduled priority queue
    scheduled_wo_ids = {e.work_order_id for e in entries}
    all_active = WorkOrder.query.filter(
        WorkOrder.location_id == location_id,
        WorkOrder.is_closed == False,
        WorkOrder.status.notin_(['Completed', 'Cancelled', 'Billed', 'To Arrive']),
    ).all()
    unscheduled = [w for w in all_active if w.id not in scheduled_wo_ids]
    unscheduled.sort(key=lambda w: (w.rev_per_hour, w.days_on_site or 0), reverse=True)
    unscheduled_data = [
        {'wo': wo, 'car': wo.car, 'warnings': get_scheduling_warnings(wo)}
        for wo in unscheduled
    ]

    # Monthly target
    target = MonthlyTarget.query.filter_by(
        location_id=location_id, month=month, year=year
    ).first()

    # Stats
    completed_count = sum(1 for e in entries if e.schedule_status == 'Completed')
    total_revenue   = 0
    for e in entries:
        wo = WorkOrder.query.get(e.work_order_id)
        if wo:
            total_revenue += wo.estimated_revenue or 0

    # Service mix
    mix = {}
    for e in entries:
        wo = WorkOrder.query.get(e.work_order_id)
        if wo and wo.car and wo.car.customer:
            cname      = wo.car.customer.name
            mix[cname] = mix.get(cname, 0) + 1

    # Gantt data
    days_in_month = calendar.monthrange(year, month)[1]
    month_start   = date(year, month, 1)
    month_end     = date(year, month, days_in_month)

    gantt_data = []
    for entry in entries:
        wo = WorkOrder.query.get(entry.work_order_id)
        if not wo:
            continue
        start = wo.repair_start_date or entry.target_date or month_start
        end   = wo.projected_outbound or month_end
        start_pct = max(0, (start - month_start).days / days_in_month * 100)
        end_pct   = min(100, (end   - month_start).days / days_in_month * 100)
        width_pct = max(2, end_pct - start_pct)
        gantt_data.append({
            'entry':     entry,
            'wo':        wo,
            'car':       wo.car,
            'start_pct': round(start_pct, 1),
            'width_pct': round(width_pct, 1),
            'warnings':  get_scheduling_warnings(wo),
        })

    # Today marker position
    today_pct = max(0, min(100, (date.today() - month_start).days / days_in_month * 100))

    return render_template('schedule.html',
        user=user,
        location=location,
        locations=locations,
        weeks=weeks,
        month=month,
        year=year,
        view=view,
        entries=entries,
        unscheduled_data=unscheduled_data,
        target=target,
        completed_count=completed_count,
        total_revenue=total_revenue,
        mix=mix,
        gantt_data=gantt_data,
        days_in_month=days_in_month,
        month_start=month_start,
        today_pct=round(today_pct, 1),
        get_status_color=get_status_color,
        get_scheduling_warnings=get_scheduling_warnings,
    )


# ─────────────────────────────────────────────
# ADD TO MONTHLY SCHEDULE
# ─────────────────────────────────────────────

@main.route('/schedule/add', methods=['POST'])
def add_to_schedule():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.form.get('location_id', type=int)
    wo_id       = request.form.get('wo_id', type=int)
    wo          = WorkOrder.query.get(wo_id)

    entry = ScheduleEntry(
        work_order_id=wo_id,
        location_id=location_id,
        month=int(request.form.get('month')),
        year=int(request.form.get('year')),
        target_week=int(request.form.get('target_week', 1)),
        target_date=_parse_date(request.form.get('target_date')),
        schedule_status='Scheduled',
        rev_per_hour=wo.rev_per_hour if wo else 0,
        priority_score=wo.rev_per_hour if wo else 0,
        notes=request.form.get('notes', '').strip(),
        scheduled_by=user.name,
    )
    db.session.add(entry)

    if wo:
        target_date_val = _parse_date(request.form.get('target_date'))
        if target_date_val:
            wo.projected_outbound = target_date_val
        auto_calculate_status(wo)
        wo.updated_by = user.name

    db.session.commit()
    flash('Car added to schedule.', 'success')

    return redirect(url_for('main.schedule',
        location_id=location_id,
        month=entry.month,
        year=entry.year,
    ))


# ─────────────────────────────────────────────
# SET MONTHLY TARGET
# ─────────────────────────────────────────────

@main.route('/schedule/set-target', methods=['POST'])
def set_monthly_target():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id    = request.form.get('location_id', type=int)
    month          = request.form.get('month', type=int)
    year           = request.form.get('year', type=int)
    car_target     = request.form.get('car_target', type=int, default=40)
    revenue_target = float(request.form.get('revenue_target') or 0)

    target = MonthlyTarget.query.filter_by(
        location_id=location_id, month=month, year=year
    ).first()
    if target:
        target.car_target     = car_target
        target.revenue_target = revenue_target
    else:
        target = MonthlyTarget(
            location_id=location_id,
            month=month,
            year=year,
            car_target=car_target,
            revenue_target=revenue_target,
            created_by=user.name,
        )
        db.session.add(target)

    db.session.commit()
    flash(f'Monthly target updated: {car_target} cars / ${revenue_target:,.0f}', 'success')
    return redirect(url_for('main.schedule',
        location_id=location_id, month=month, year=year))


# ─────────────────────────────────────────────
# REMOVE FROM SCHEDULE
# ─────────────────────────────────────────────

@main.route('/schedule/remove/<int:entry_id>', methods=['POST'])
def remove_from_schedule(entry_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    entry       = ScheduleEntry.query.get_or_404(entry_id)
    wo          = WorkOrder.query.get(entry.work_order_id)
    location_id = entry.location_id
    month       = entry.month
    year        = entry.year

    db.session.delete(entry)

    if wo:
        auto_calculate_status(wo)
        wo.updated_by = user.name
        log_activity(wo.id, user.name, user.role, 'Removed from schedule')

    db.session.commit()
    flash('Car removed from schedule.', 'success')
    return redirect(url_for('main.schedule',
        location_id=location_id, month=month, year=year))


# ─────────────────────────────────────────────
# UPDATE SCHEDULE ENTRY STATUS
# ─────────────────────────────────────────────

@main.route('/schedule/update/<int:entry_id>', methods=['POST'])
def update_schedule_entry(entry_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    entry                  = ScheduleEntry.query.get_or_404(entry_id)
    old_status             = entry.schedule_status
    entry.schedule_status  = request.form.get('schedule_status', entry.schedule_status)
    entry.target_week      = request.form.get('target_week', type=int, default=entry.target_week)
    entry.notes            = request.form.get('notes', '').strip()

    wo = WorkOrder.query.get(entry.work_order_id)
    if wo and entry.schedule_status != old_status:
        if entry.schedule_status == 'Completed':
            if not wo.repair_completed_date:
                wo.repair_completed_date = date.today()
        auto_calculate_status(wo)
        wo.updated_by = user.name

    db.session.commit()

    # Handle remove action (legacy _remove field from schedule.html)
    if request.form.get('_remove') == '1':
        # Re-load to do proper delete via remove_from_schedule logic
        return redirect(url_for('main.schedule',
            location_id=entry.location_id,
            month=entry.month,
            year=entry.year,
        ))

    return redirect(url_for('main.schedule',
        location_id=entry.location_id,
        month=entry.month,
        year=entry.year,
    ))


# ─────────────────────────────────────────────
# API — search cars (for quick lookup)
# ─────────────────────────────────────────────

@main.route('/api/search-cars')
def search_cars():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    cars = Car.query.filter(
        db.or_(
            Car.car_mark.ilike(f'%{q}%'),
            Car.car_number.ilike(f'%{q}%'),
        )
    ).limit(10).all()
    results = []
    for car in cars:
        wo = car.active_work_order
        results.append({
            'id':       car.id,
            'full_id':  car.full_id,
            'customer': car.customer.name if car.customer else '',
            'status':   wo.status if wo else 'No active WO',
            'wo_id':    wo.id if wo else None,
        })
    return jsonify(results)


# ─────────────────────────────────────────────
# GANTT SCHEDULE
# ─────────────────────────────────────────────

@main.route('/schedule/gantt')
def schedule_gantt():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id

    location  = Location.query.get(location_id)
    locations = Location.query.all()

    week_offset = request.args.get('week_offset', type=int, default=0)

    today       = date.today()
    month_start = date(today.year, today.month, 1)
    view_start  = month_start + timedelta(weeks=week_offset)
    view_end    = view_start + timedelta(days=27)   # always 28 days / 4 weeks

    date_range = []
    d = view_start
    while d <= view_end:
        date_range.append(d)
        d += timedelta(days=1)

    all_spots   = Spot.query.filter_by(location_id=location_id, is_active=True).all()
    gantt_spots = sorted(
        [s for s in all_spots if is_gantt_spot(s.name, location.name)],
        key=lambda s: s.name,
    )

    marks = GanttMark.query.filter(
        GanttMark.location_id == location_id,
        GanttMark.mark_date   >= view_start,
        GanttMark.mark_date   <= view_end,
    ).all()

    mark_map = {}
    for mark in marks:
        mark_map.setdefault(mark.spot_id, {})[mark.mark_date] = mark

    scheduled_entries = ScheduleEntry.query.filter_by(
        location_id=location_id,
    ).filter(
        ScheduleEntry.schedule_status.notin_(['Completed', 'Pushed to Next Month'])
    ).order_by(ScheduleEntry.rev_per_hour.desc()).all()

    scheduled_wos = []
    for entry in scheduled_entries:
        wo = WorkOrder.query.get(entry.work_order_id)
        if wo and not wo.is_closed:
            scheduled_wos.append({
                'entry':    entry,
                'wo':       wo,
                'car':      wo.car,
                'warnings': get_scheduling_warnings(wo),
            })

    capacity   = LOCATION_CAPACITY.get(location.name, 500)
    week_usage = {}
    for mark in marks:
        wk  = (mark.mark_date - view_start).days // 7
        wo  = WorkOrder.query.get(mark.work_order_id)
        if wo:
            week_usage.setdefault(wk, {})
            week_usage[wk].setdefault(wo.id, wo.estimated_hours or 0)

    week_hours = {wk: sum(hrs.values()) for wk, hrs in week_usage.items()}

    return render_template('schedule_gantt.html',
        user=user,
        location=location,
        locations=locations,
        gantt_spots=gantt_spots,
        date_range=date_range,
        mark_map=mark_map,
        scheduled_wos=scheduled_wos,
        week_offset=week_offset,
        view_start=view_start,
        view_end=view_end,
        week_hours=week_hours,
        capacity=capacity,
        today=today,
        get_status_color=get_status_color,
    )


@main.route('/schedule/gantt/mark', methods=['POST'])
def gantt_mark():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Not logged in'}), 401

    data         = request.json or {}
    spot_id      = data.get('spot_id')
    wo_id        = data.get('wo_id')
    mark_date_str= data.get('mark_date')
    location_id  = data.get('location_id')

    if not all([spot_id, wo_id, mark_date_str, location_id]):
        return jsonify({'error': 'Missing fields'}), 400

    try:
        mark_date = datetime.strptime(mark_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date'}), 400

    existing = GanttMark.query.filter_by(spot_id=spot_id, mark_date=mark_date).first()

    if existing:
        if existing.work_order_id == wo_id:
            db.session.delete(existing)
            db.session.commit()
            return jsonify({'action': 'removed', 'spot_id': spot_id, 'date': mark_date_str})
        conflict_wo  = WorkOrder.query.get(existing.work_order_id)
        conflict_car = conflict_wo.car.full_id if conflict_wo else 'Unknown'
        return jsonify({
            'error':        'conflict',
            'message':      f'Spot already booked by {conflict_car} on this date.',
            'existing_car': conflict_car,
        }), 409

    mark = GanttMark(
        spot_id=spot_id, work_order_id=wo_id,
        location_id=location_id, mark_date=mark_date,
        created_by=user.name,
    )
    db.session.add(mark)

    wo = WorkOrder.query.get(wo_id)
    if wo:
        all_dates = [m.mark_date for m in GanttMark.query.filter_by(work_order_id=wo_id).all()]
        all_dates.append(mark_date)
        sa = SpotAssignment.query.filter_by(work_order_id=wo_id, is_active=True).first()
        if sa:
            sa.scheduled_start_date = min(all_dates)
            sa.scheduled_end_date   = max(all_dates)

    db.session.commit()
    return jsonify({
        'action':  'added',
        'spot_id': spot_id,
        'wo_id':   wo_id,
        'date':    mark_date_str,
        'car_id':  wo.car.full_id if wo else '',
        'color':   '#2563eb',
    })


@main.route('/schedule/gantt/clear-car', methods=['POST'])
def gantt_clear_car():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Not logged in'}), 401
    wo_id = (request.json or {}).get('wo_id')
    GanttMark.query.filter_by(work_order_id=wo_id).delete()
    db.session.commit()
    return jsonify({'action': 'cleared', 'wo_id': wo_id})


# ─────────────────────────────────────────────
# HISTORY / ARCHIVE
# ─────────────────────────────────────────────

@main.route('/history')
def history():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))

    location_id = request.args.get('location_id', type=int) \
                  or user.location_id \
                  or Location.query.first().id
    locations = Location.query.all()
    location  = Location.query.get(location_id)

    period   = request.args.get('period', 'month')
    status_f = request.args.get('status', '')
    search   = request.args.get('search', '').strip()

    today = date.today()
    if period == 'week':
        period_start = today - timedelta(days=7)
    elif period == 'month':
        period_start = date(today.year, today.month, 1)
    elif period == 'quarter':
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        period_start = date(today.year, quarter_month, 1)
    elif period == 'year':
        period_start = date(today.year, 1, 1)
    else:
        period_start = date(2000, 1, 1)

    query = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == True)

    if period != 'all':
        query = query.filter(WorkOrder.billed_out >= period_start)

    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)

    if status_f:
        query = query.filter(WorkOrder.status == status_f)

    if search:
        query = query.filter(db.or_(
            Car.car_mark.ilike(f'%{search}%'),
            Car.car_number.ilike(f'%{search}%'),
        ))

    work_orders = query.order_by(WorkOrder.billed_out.desc().nullslast()).all()

    total_cars = len(work_orders)
    total_revenue = sum(wo.estimated_revenue or 0 for wo in work_orders)

    valid_turnaround = [
        (wo.billed_out - wo.date_arrived).days
        for wo in work_orders
        if wo.billed_out and wo.date_arrived
    ]
    avg_turnaround = round(
        sum(valid_turnaround) / len(valid_turnaround)
    ) if valid_turnaround else 0

    valid_rph  = [wo.rev_per_hour for wo in work_orders if wo.rev_per_hour > 0]
    avg_rev_hr = round(sum(valid_rph) / len(valid_rph), 2) if valid_rph else 0

    history_stats = {
        'count':      total_cars,
        'total_rev':  total_revenue,
        'avg_days':   avg_turnaround,
        'avg_rev_hr': avg_rev_hr,
    }

    # Most recent YardCar per WO (needed for Undo Bill Out button)
    yc_map = {}
    for wo in work_orders:
        yc = YardCar.query.filter_by(work_order_id=wo.id) \
                 .order_by(YardCar.entered_at.desc()).first()
        if yc:
            yc_map[wo.id] = yc

    return render_template('history.html',
        user=user,
        location=location,
        locations=locations,
        work_orders=work_orders,
        history_stats=history_stats,
        yc_map=yc_map,
        period=period,
        status_f=status_f,
        search=search,
        today=today,
        get_status_color=get_status_color,
    )


# ─────────────────────────────────────────────
# REPORTS PAGE
# ─────────────────────────────────────────────

@main.route('/reports')
def reports():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    locations = Location.query.all()
    today = date.today()
    return render_template('reports.html',
        user=user,
        locations=locations,
        today=today,
        current_month=today.month,
        current_year=today.year,
    )


# ─────────────────────────────────────────────
# REPORT 1 — MONTHLY SUMMARY
# ─────────────────────────────────────────────

@main.route('/reports/monthly-summary/excel')
def report_monthly_summary_excel():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not OPENPYXL_AVAILABLE:
        flash('Excel library not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('main.reports'))

    month       = request.args.get('month', type=int, default=date.today().month)
    year        = request.args.get('year',  type=int, default=date.today().year)
    location_id = request.args.get('location_id', type=int)

    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])

    query = WorkOrder.query.filter(
        WorkOrder.billed_out >= month_start,
        WorkOrder.billed_out <= month_end,
        WorkOrder.is_closed == True,
    )
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    wos = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Monthly Summary'
    styles   = style_excel_workbook(wb, '')

    month_name = date(year, month, 1).strftime('%B %Y')
    loc_label  = Location.query.get(location_id).name if location_id else 'All Locations'
    _xl_title_block(ws, 'KRONIX — Monthly Summary Report',
                    f'{month_name} · {loc_label}', col_span=10)

    total_rev = sum(w.estimated_revenue or 0 for w in wos)
    valid     = [(w.billed_out - w.date_arrived).days
                 for w in wos if w.billed_out and w.date_arrived]
    avg_days  = round(sum(valid) / len(valid)) if valid else 0

    for col, label in enumerate(['Total Cars', 'Total Revenue', 'Avg Turnaround'], 1):
        ws.cell(row=4, column=col*2-1, value=label).font = Font(bold=True, color='64748B', name='Calibri')
    ws.cell(row=4, column=2, value=len(wos)).font  = Font(bold=True, size=12, name='Calibri')
    ws.cell(row=4, column=4, value=f'${total_rev:,.0f}').font = Font(bold=True, size=12, name='Calibri')
    ws.cell(row=4, column=6, value=f'{avg_days} days').font   = Font(bold=True, size=12, name='Calibri')
    ws.row_dimensions[4].height = 20

    headers = ['Car ID', 'Customer', 'Location', 'Purpose', 'Arrived',
               'Billed Out', 'Days on Site', 'Est. Revenue', 'Rev/Hr', 'Status']
    _apply_xl_header_row(ws, headers, 6, styles)

    for i, wo in enumerate(wos):
        loc  = Location.query.get(wo.location_id)
        days = (wo.billed_out - wo.date_arrived).days if (wo.billed_out and wo.date_arrived) else ''
        _xl_data_row(ws, 7 + i, [
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            loc.name if loc else '',
            wo.purpose or '',
            wo.date_arrived.strftime('%Y-%m-%d') if wo.date_arrived else '',
            wo.billed_out.strftime('%Y-%m-%d')   if wo.billed_out   else '',
            days,
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
            f'${wo.rev_per_hour}/hr'         if wo.rev_per_hour      else '',
            wo.status,
        ], styles)

    for col, w in enumerate([15, 18, 12, 18, 13, 13, 13, 14, 12, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return excel_response(wb, f'kronix-monthly-summary-{year}-{month:02d}')


@main.route('/reports/monthly-summary/pdf')
def report_monthly_summary_pdf():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not REPORTLAB_AVAILABLE:
        flash('PDF library not installed. Run: pip install reportlab', 'error')
        return redirect(url_for('main.reports'))

    month       = request.args.get('month', type=int, default=date.today().month)
    year        = request.args.get('year',  type=int, default=date.today().year)
    location_id = request.args.get('location_id', type=int)

    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])

    query = WorkOrder.query.filter(
        WorkOrder.billed_out >= month_start,
        WorkOrder.billed_out <= month_end,
        WorkOrder.is_closed == True,
    )
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    wos = query.all()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []

    month_name = date(year, month, 1).strftime('%B %Y')
    loc_label  = Location.query.get(location_id).name if location_id else 'All Locations'
    _pdf_title_elements('KRONIX — Monthly Summary Report',
                        f'{month_name} · {loc_label} · Generated {date.today().strftime("%B %d, %Y")}',
                        elements)

    total_rev = sum(w.estimated_revenue or 0 for w in wos)
    valid     = [(w.billed_out - w.date_arrived).days for w in wos if w.billed_out and w.date_arrived]
    avg_days  = round(sum(valid) / len(valid)) if valid else 0

    stats_data = [
        ['Total Cars', 'Total Revenue', 'Avg Turnaround', 'Report Period'],
        [str(len(wos)), f'${total_rev:,.0f}', f'{avg_days} days', month_name],
    ]
    stats_tbl = Table(stats_data, colWidths=[2.5*inch]*4)
    stats_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#EFF6FF')),
        ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,1), (-1,1), 14),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWHEIGHT',  (0,0), (-1,-1), 28),
    ]))
    elements.append(stats_tbl)
    elements.append(Spacer(1, 0.2*inch))

    col_headers = ['Car ID', 'Customer', 'Location', 'Purpose',
                   'Arrived', 'Billed Out', 'Days', 'Revenue', 'Status']
    table_data  = [col_headers]
    for wo in wos:
        loc  = Location.query.get(wo.location_id)
        days = (wo.billed_out - wo.date_arrived).days if (wo.billed_out and wo.date_arrived) else ''
        table_data.append([
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            loc.name if loc else '',
            wo.purpose or '',
            wo.date_arrived.strftime('%m/%d/%y') if wo.date_arrived else '',
            wo.billed_out.strftime('%m/%d/%y')   if wo.billed_out   else '',
            str(days),
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
            wo.status,
        ])

    col_widths = [1.1*inch, 1.3*inch, 0.9*inch, 1.3*inch, 0.9*inch,
                  0.9*inch, 0.5*inch, 0.9*inch, 1.2*inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(table_data)))
    elements.append(tbl)
    _pdf_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    return pdf_response(buffer, f'kronix-monthly-summary-{year}-{month:02d}')


# ─────────────────────────────────────────────
# REPORT 2 — CURRENT WIP
# ─────────────────────────────────────────────

@main.route('/reports/wip/excel')
def report_wip_excel():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not OPENPYXL_AVAILABLE:
        flash('Excel library not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)
    query = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == False)
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    wos = query.all()
    wos.sort(key=lambda w: (w.date_arrived or date.min), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Current WIP'
    styles   = style_excel_workbook(wb, '')

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _xl_title_block(ws, 'KRONIX — Current WIP Report',
                    f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', col_span=11)

    headers = ['Car ID', 'Customer', 'Location', 'Purpose', 'Car Status',
               'Arrived', 'Days on Site', 'Proj. Outbound', 'Overdue', 'Parts Status', 'Est. Revenue']
    _apply_xl_header_row(ws, headers, 4, styles)

    for i, wo in enumerate(wos):
        loc   = Location.query.get(wo.location_id)
        today = date.today()
        days  = (today - wo.date_arrived).days if wo.date_arrived else ''
        overdue = 'Yes' if wo.is_overdue else 'No'
        row_data = [
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            loc.name if loc else '',
            wo.purpose or '',
            wo.status,
            wo.date_arrived.strftime('%Y-%m-%d')       if wo.date_arrived      else '',
            days,
            wo.projected_outbound.strftime('%Y-%m-%d') if wo.projected_outbound else '',
            overdue,
            wo.parts_status or '',
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ]
        _xl_data_row(ws, 5 + i, row_data, styles)
        if overdue == 'Yes':
            red_fill = PatternFill('solid', fgColor='FEF2F2')
            for col in range(1, 12):
                ws.cell(row=5 + i, column=col).fill = red_fill

    for col, w in enumerate([15, 18, 12, 16, 18, 13, 12, 14, 9, 16, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return excel_response(wb, f'kronix-wip-{date.today().strftime("%Y-%m-%d")}')


@main.route('/reports/wip/pdf')
def report_wip_pdf():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not REPORTLAB_AVAILABLE:
        flash('PDF library not installed. Run: pip install reportlab', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)
    query = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == False)
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    wos = query.all()
    wos.sort(key=lambda w: (w.date_arrived or date.min), reverse=True)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _pdf_title_elements('KRONIX — Current WIP Report',
                        f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', elements)

    col_headers = ['Car ID', 'Customer', 'Location', 'Purpose', 'Status',
                   'Arrived', 'Days', 'Proj. Out', 'Overdue', 'Parts Status', 'Revenue']
    table_data  = [col_headers]
    overdue_rows = []
    today = date.today()

    for wo in wos:
        loc  = Location.query.get(wo.location_id)
        days = (today - wo.date_arrived).days if wo.date_arrived else ''
        row_idx = len(table_data)
        if wo.is_overdue:
            overdue_rows.append(row_idx)
        table_data.append([
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            loc.name if loc else '',
            wo.purpose or '',
            wo.status,
            wo.date_arrived.strftime('%m/%d/%y')       if wo.date_arrived      else '',
            str(days),
            wo.projected_outbound.strftime('%m/%d/%y') if wo.projected_outbound else '',
            'Yes' if wo.is_overdue else 'No',
            wo.parts_status or '',
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ])

    col_widths = [1.0*inch, 1.1*inch, 0.8*inch, 1.0*inch, 1.1*inch,
                  0.8*inch, 0.45*inch, 0.8*inch, 0.55*inch, 0.9*inch, 0.85*inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(table_data), overdue_rows=overdue_rows))
    elements.append(tbl)
    _pdf_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    return pdf_response(buffer, f'kronix-wip-{date.today().strftime("%Y-%m-%d")}')


# ─────────────────────────────────────────────
# REPORT 3 — SHOP FLOOR UTILIZATION
# ─────────────────────────────────────────────

@main.route('/reports/shop-utilization/excel')
def report_shop_utilization_excel():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not OPENPYXL_AVAILABLE:
        flash('Excel library not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)

    spot_q = Spot.query
    if location_id:
        spot_q = spot_q.filter_by(location_id=location_id)
    spots = spot_q.all()
    spot_ids = [s.id for s in spots]

    assignments = SpotAssignment.query.filter(
        SpotAssignment.spot_id.in_(spot_ids)
    ).order_by(SpotAssignment.assigned_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Shop Utilization'
    styles   = style_excel_workbook(wb, '')

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _xl_title_block(ws, 'KRONIX — Shop Floor Utilization Report',
                    f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', col_span=12)

    occupied  = len({a.spot_id for a in assignments if a.is_active})
    total_est = sum(a.estimated_hours or 0 for a in assignments)
    total_act = sum(a.actual_hours    or 0 for a in assignments)
    util_pct  = round(occupied / len(spots) * 100) if spots else 0

    for col, (lbl, val) in enumerate([
        ('Total Spots', len(spots)),
        ('Occupied',    occupied),
        ('Utilization', f'{util_pct}%'),
        ('Total Est. Hrs', f'{total_est:.1f}'),
        ('Total Act. Hrs', f'{total_act:.1f}'),
    ], 1):
        ws.cell(row=4, column=col*2-1, value=lbl).font = Font(bold=True, color='64748B', name='Calibri')
        ws.cell(row=4, column=col*2,   value=val).font = Font(bold=True, size=11, name='Calibri')
    ws.row_dimensions[4].height = 20

    headers = ['Spot', 'Car ID', 'Customer', 'SOW Type', 'Est. Hours', 'Actual Hours',
               '% Complete', 'Status', 'Assigned Date', 'Target Date', 'Days in Spot', 'Active']
    _apply_xl_header_row(ws, headers, 6, styles)

    for i, a in enumerate(assignments):
        wo  = a.work_order
        car = wo.car if wo else None
        spot = Spot.query.get(a.spot_id)
        _xl_data_row(ws, 7 + i, [
            spot.name if spot else '',
            car.full_id if car else '',
            car.customer.name if (car and car.customer) else '',
            a.sow_type or '',
            a.estimated_hours or 0,
            a.actual_hours    or 0,
            f'{a.pct_complete}%',
            a.status or '',
            a.assigned_at.strftime('%Y-%m-%d') if a.assigned_at else '',
            a.target_date.strftime('%Y-%m-%d') if a.target_date else '',
            a.days_in_spot,
            'Yes' if a.is_active else 'No',
        ], styles)

    for col, w in enumerate([14, 14, 16, 14, 11, 12, 11, 14, 14, 13, 12, 8], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return excel_response(wb, f'kronix-shop-utilization-{date.today().strftime("%Y-%m-%d")}')


@main.route('/reports/shop-utilization/pdf')
def report_shop_utilization_pdf():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not REPORTLAB_AVAILABLE:
        flash('PDF library not installed. Run: pip install reportlab', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)

    spot_q = Spot.query
    if location_id:
        spot_q = spot_q.filter_by(location_id=location_id)
    spots    = spot_q.all()
    spot_ids = [s.id for s in spots]

    assignments = SpotAssignment.query.filter(
        SpotAssignment.spot_id.in_(spot_ids)
    ).order_by(SpotAssignment.is_active.desc(), SpotAssignment.assigned_at.desc()).all()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _pdf_title_elements('KRONIX — Shop Floor Utilization Report',
                        f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', elements)

    occupied  = len({a.spot_id for a in assignments if a.is_active})
    total_est = sum(a.estimated_hours or 0 for a in assignments)
    total_act = sum(a.actual_hours    or 0 for a in assignments)
    util_pct  = round(occupied / len(spots) * 100) if spots else 0

    stats_data = [
        ['Total Spots', 'Occupied', 'Utilization', 'Total Est. Hrs', 'Total Act. Hrs'],
        [str(len(spots)), str(occupied), f'{util_pct}%', f'{total_est:.1f}', f'{total_act:.1f}'],
    ]
    stats_tbl = Table(stats_data, colWidths=[2.0*inch]*5)
    stats_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#EFF6FF')),
        ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,1), (-1,1), 13),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWHEIGHT',  (0,0), (-1,-1), 26),
    ]))
    elements.append(stats_tbl)
    elements.append(Spacer(1, 0.2*inch))

    col_headers = ['Spot', 'Car ID', 'Customer', 'SOW Type', 'Est. Hrs',
                   'Act. Hrs', '% Done', 'Status', 'Assigned', 'Target', 'Days', 'Active']
    table_data  = [col_headers]
    for a in assignments:
        wo   = a.work_order
        car  = wo.car if wo else None
        spot = Spot.query.get(a.spot_id)
        table_data.append([
            spot.name if spot else '',
            car.full_id if car else '',
            car.customer.name if (car and car.customer) else '',
            a.sow_type or '',
            f'{a.estimated_hours or 0:.1f}',
            f'{a.actual_hours or 0:.1f}',
            f'{a.pct_complete}%',
            a.status or '',
            a.assigned_at.strftime('%m/%d/%y') if a.assigned_at else '',
            a.target_date.strftime('%m/%d/%y') if a.target_date else '',
            str(a.days_in_spot),
            'Y' if a.is_active else 'N',
        ])

    col_widths = [1.0*inch, 1.0*inch, 1.1*inch, 1.0*inch, 0.65*inch,
                  0.65*inch, 0.55*inch, 1.0*inch, 0.75*inch, 0.75*inch, 0.5*inch, 0.4*inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(table_data)))
    elements.append(tbl)
    _pdf_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    return pdf_response(buffer, f'kronix-shop-utilization-{date.today().strftime("%Y-%m-%d")}')


# ─────────────────────────────────────────────
# REPORT 4 — YARD STATUS
# ─────────────────────────────────────────────

@main.route('/reports/yard-status/excel')
def report_yard_status_excel():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not OPENPYXL_AVAILABLE:
        flash('Excel library not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)

    query = YardCar.query.join(WorkOrder).filter(YardCar.is_active == True)
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    yard_cars = query.all()
    yard_cars.sort(key=lambda yc: (yc.yard_status or '', yc.entered_at or datetime.min), reverse=False)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Yard Status'
    styles   = style_excel_workbook(wb, '')

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _xl_title_block(ws, 'KRONIX — Yard Status Report',
                    f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', col_span=10)

    headers = ['Car ID', 'Customer', 'Commodity', 'Car Status', 'Yard Status',
               'Return Location', 'Arrival Date', 'Days in Yard', 'Proj. Outbound', 'Notes']
    _apply_xl_header_row(ws, headers, 4, styles)

    current_status = None
    row = 5
    for yc in yard_cars:
        wo  = yc.work_order
        car = wo.car if wo else None

        if yc.yard_status != current_status:
            current_status = yc.yard_status
            ws.merge_cells(f'A{row}:J{row}')
            cell = ws.cell(row=row, column=1, value=current_status or 'Unknown')
            cell.fill = PatternFill('solid', fgColor='1E293B')
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[row].height = 18
            row += 1

        today  = date.today()
        days   = (today - yc.entered_at.date()).days if yc.entered_at else ''
        _xl_data_row(ws, row, [
            car.full_id if car else '',
            car.customer.name if (car and car.customer) else '',
            car.commodity     if car                    else '',
            wo.status         if wo                     else '',
            yc.yard_status    or '',
            yc.return_location or '',
            yc.entered_at.strftime('%Y-%m-%d')          if yc.entered_at           else '',
            days,
            wo.projected_outbound.strftime('%Y-%m-%d')  if (wo and wo.projected_outbound) else '',
            yc.comments or '',
        ], styles)
        row += 1

    for col, w in enumerate([14, 16, 12, 16, 16, 16, 13, 12, 14, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return excel_response(wb, f'kronix-yard-status-{date.today().strftime("%Y-%m-%d")}')


@main.route('/reports/yard-status/pdf')
def report_yard_status_pdf():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not REPORTLAB_AVAILABLE:
        flash('PDF library not installed. Run: pip install reportlab', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)

    query = YardCar.query.join(WorkOrder).filter(YardCar.is_active == True)
    if location_id:
        query = query.filter(WorkOrder.location_id == location_id)
    yard_cars = query.all()
    yard_cars.sort(key=lambda yc: (yc.yard_status or '', yc.entered_at or datetime.min))

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _pdf_title_elements('KRONIX — Yard Status Report',
                        f'{loc_label} · As of {date.today().strftime("%B %d, %Y")}', elements)

    col_headers = ['Car ID', 'Customer', 'Commodity', 'Car Status', 'Yard Status',
                   'Return Location', 'Arrival Date', 'Days', 'Proj. Outbound', 'Notes']
    table_data  = [col_headers]
    today = date.today()

    for yc in yard_cars:
        wo  = yc.work_order
        car = wo.car if wo else None
        days = (today - yc.entered_at.date()).days if yc.entered_at else ''
        table_data.append([
            car.full_id if car else '',
            car.customer.name if (car and car.customer) else '',
            car.commodity     if car                    else '',
            wo.status         if wo                     else '',
            yc.yard_status    or '',
            yc.return_location or '',
            yc.entered_at.strftime('%m/%d/%y')         if yc.entered_at                else '',
            str(days),
            wo.projected_outbound.strftime('%m/%d/%y') if (wo and wo.projected_outbound) else '',
            (yc.comments or '')[:40],
        ])

    col_widths = [1.0*inch, 1.1*inch, 0.85*inch, 1.1*inch, 1.0*inch,
                  1.1*inch, 0.8*inch, 0.4*inch, 0.85*inch, 1.2*inch]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(table_data)))
    elements.append(tbl)
    _pdf_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    return pdf_response(buffer, f'kronix-yard-status-{date.today().strftime("%Y-%m-%d")}')


# ─────────────────────────────────────────────
# REPORT 5 — PARTS & DELAYS
# ─────────────────────────────────────────────

@main.route('/reports/parts-delays/excel')
def report_parts_delays_excel():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not OPENPYXL_AVAILABLE:
        flash('Excel library not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)
    today       = date.today()

    def _base_q():
        q = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == False)
        if location_id:
            q = q.filter(WorkOrder.location_id == location_id)
        return q

    # Section A — Waiting Parts
    waiting_parts = _base_q().filter(
        WorkOrder.parts_status.in_(['Ordered', 'Waiting ETA', 'Partially Available'])
    ).all()

    # Section B — No Projected Outbound
    no_proj = _base_q().filter(
        WorkOrder.projected_outbound == None,
        WorkOrder.work_order_created != None,
    ).all()

    # Section C — Overdue
    overdue_wos = _base_q().filter(
        WorkOrder.projected_outbound < today,
    ).all()

    wb = Workbook()

    # ── Sheet A ──────────────────────────────
    ws_a = wb.active
    ws_a.title = 'Waiting Parts'
    styles     = style_excel_workbook(wb, '')
    loc_label  = Location.query.get(location_id).name if location_id else 'All Locations'

    _xl_title_block(ws_a, 'KRONIX — Parts & Delays: Waiting Parts',
                    f'{loc_label} · {today.strftime("%B %d, %Y")}', col_span=7)
    headers_a = ['Car ID', 'Customer', 'Parts Status', 'Parts ETA', 'Parts Notes', 'Days Waiting', 'Car Status']
    _apply_xl_header_row(ws_a, headers_a, 4, styles)
    for i, wo in enumerate(waiting_parts):
        days_w = (today - wo.estimate_approved).days if wo.estimate_approved else ''
        _xl_data_row(ws_a, 5 + i, [
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.parts_status or '',
            wo.parts_eta.strftime('%Y-%m-%d') if wo.parts_eta else '',
            wo.parts_notes or '',
            days_w,
            wo.status,
        ], styles)
    for col, w in enumerate([14, 18, 18, 13, 26, 13, 18], 1):
        ws_a.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet B ──────────────────────────────
    ws_b = wb.create_sheet('No Projected Outbound')
    _xl_title_block(ws_b, 'KRONIX — Parts & Delays: No Projected Outbound',
                    f'{loc_label} · {today.strftime("%B %d, %Y")}', col_span=6)
    headers_b = ['Car ID', 'Customer', 'Status', 'Work Order Created', 'Days Since WO', 'Est. Revenue']
    _apply_xl_header_row(ws_b, headers_b, 4, styles)
    for i, wo in enumerate(no_proj):
        days_wo = (today - wo.work_order_created).days if wo.work_order_created else ''
        _xl_data_row(ws_b, 5 + i, [
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.status,
            wo.work_order_created.strftime('%Y-%m-%d') if wo.work_order_created else '',
            days_wo,
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ], styles)
    for col, w in enumerate([14, 18, 18, 18, 14, 14], 1):
        ws_b.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet C ──────────────────────────────
    ws_c = wb.create_sheet('Overdue Cars')
    _xl_title_block(ws_c, 'KRONIX — Parts & Delays: Overdue Cars',
                    f'{loc_label} · {today.strftime("%B %d, %Y")}', col_span=6)
    headers_c = ['Car ID', 'Customer', 'Proj. Outbound', 'Days Overdue', 'Status', 'Est. Revenue']
    _apply_xl_header_row(ws_c, headers_c, 4, styles)
    for i, wo in enumerate(overdue_wos):
        _xl_data_row(ws_c, 5 + i, [
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.projected_outbound.strftime('%Y-%m-%d') if wo.projected_outbound else '',
            wo.days_overdue,
            wo.status,
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ], styles)
    for col, w in enumerate([14, 18, 15, 13, 18, 14], 1):
        ws_c.column_dimensions[get_column_letter(col)].width = w

    return excel_response(wb, f'kronix-parts-delays-{today.strftime("%Y-%m-%d")}')


@main.route('/reports/parts-delays/pdf')
def report_parts_delays_pdf():
    user = get_current_user()
    if not user:
        return redirect(url_for('main.login'))
    if not REPORTLAB_AVAILABLE:
        flash('PDF library not installed. Run: pip install reportlab', 'error')
        return redirect(url_for('main.reports'))

    location_id = request.args.get('location_id', type=int)
    today       = date.today()

    def _base_q():
        q = WorkOrder.query.join(Car).filter(WorkOrder.is_closed == False)
        if location_id:
            q = q.filter(WorkOrder.location_id == location_id)
        return q

    waiting_parts = _base_q().filter(
        WorkOrder.parts_status.in_(['Ordered', 'Waiting ETA', 'Partially Available'])
    ).all()
    no_proj = _base_q().filter(
        WorkOrder.projected_outbound == None,
        WorkOrder.work_order_created != None,
    ).all()
    overdue_wos = _base_q().filter(
        WorkOrder.projected_outbound < today,
    ).all()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []

    loc_label = Location.query.get(location_id).name if location_id else 'All Locations'
    _pdf_title_elements('KRONIX — Parts & Delays Report',
                        f'{loc_label} · Generated {today.strftime("%B %d, %Y")}', elements)

    # Section A
    _pdf_section_header(f'A — Waiting Parts  ({len(waiting_parts)} cars)', elements)
    data_a = [['Car ID', 'Customer', 'Parts Status', 'Parts ETA', 'Parts Notes', 'Days Waiting', 'Car Status']]
    for wo in waiting_parts:
        days_w = (today - wo.estimate_approved).days if wo.estimate_approved else ''
        data_a.append([
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.parts_status or '',
            wo.parts_eta.strftime('%m/%d/%y') if wo.parts_eta else '',
            (wo.parts_notes or '')[:35],
            str(days_w),
            wo.status,
        ])
    tbl_a = Table(data_a, colWidths=[1.0*inch, 1.2*inch, 1.1*inch, 0.8*inch, 2.0*inch, 0.8*inch, 1.2*inch], repeatRows=1)
    tbl_a.setStyle(_pdf_table_style(len(data_a)))
    elements.append(tbl_a)
    elements.append(Spacer(1, 0.15*inch))

    # Section B
    _pdf_section_header(f'B — No Projected Outbound  ({len(no_proj)} cars)', elements)
    data_b = [['Car ID', 'Customer', 'Status', 'WO Created', 'Days Since WO', 'Est. Revenue']]
    for wo in no_proj:
        days_wo = (today - wo.work_order_created).days if wo.work_order_created else ''
        data_b.append([
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.status,
            wo.work_order_created.strftime('%m/%d/%y') if wo.work_order_created else '',
            str(days_wo),
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ])
    tbl_b = Table(data_b, colWidths=[1.1*inch, 1.4*inch, 1.4*inch, 1.0*inch, 1.0*inch, 1.0*inch], repeatRows=1)
    tbl_b.setStyle(_pdf_table_style(len(data_b)))
    elements.append(tbl_b)
    elements.append(Spacer(1, 0.15*inch))

    # Section C
    _pdf_section_header(f'C — Overdue Cars  ({len(overdue_wos)} cars)', elements)
    data_c = [['Car ID', 'Customer', 'Proj. Outbound', 'Days Overdue', 'Status', 'Est. Revenue']]
    for wo in overdue_wos:
        data_c.append([
            wo.car.full_id,
            wo.car.customer.name if wo.car.customer else '',
            wo.projected_outbound.strftime('%m/%d/%y') if wo.projected_outbound else '',
            str(wo.days_overdue),
            wo.status,
            f'${wo.estimated_revenue:,.0f}' if wo.estimated_revenue else '',
        ])
    tbl_c = Table(data_c, colWidths=[1.1*inch, 1.4*inch, 1.1*inch, 1.0*inch, 1.4*inch, 1.0*inch], repeatRows=1)
    tbl_c.setStyle(_pdf_table_style(len(data_c), overdue_rows=list(range(1, len(data_c)))))
    elements.append(tbl_c)
    _pdf_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    return pdf_response(buffer, f'kronix-parts-delays-{today.strftime("%Y-%m-%d")}')


# ─────────────────────────────────────────────
# UTILITY
# ─────────────────────────────────────────────

def _parse_date(value):
    """Safely parse a date string from a form field."""
    if not value or value.strip() == '':
        return None
    try:
        return datetime.strptime(value.strip(), '%Y-%m-%d').date()
    except ValueError:
        return None